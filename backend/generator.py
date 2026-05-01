import json
import os
from datetime import datetime
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    HRFlowable, PageBreak, Image
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
from docx import Document as DocxDocument
from docx.shared import Inches, Pt, RGBColor
from docx.enum.text import WD_ALIGN_PARAGRAPH
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Colores corporativos ─────────────────────────────────────────────────────
COLOR_PRIMARIO = colors.HexColor("#1a3c5e")    # Azul marino
COLOR_SECUNDARIO = colors.HexColor("#2e7d9b")   # Azul medio
COLOR_ACENTO = colors.HexColor("#f5a623")       # Dorado
COLOR_FONDO = colors.HexColor("#f0f4f8")        # Gris muy claro
COLOR_TEXTO = colors.HexColor("#1a1a2e")        # Casi negro


def _fmt_fecha_larga(fecha: datetime | None = None) -> str:
    meses = [
        "enero", "febrero", "marzo", "abril", "mayo", "junio",
        "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"
    ]
    dt = fecha or datetime.now()
    return f"{dt.day} de {meses[dt.month - 1]} de {dt.year}"


def _draw_header_footer(canvas, doc, titulo_doc: str = "Documento"):
    canvas.saveState()
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(colors.grey)
    canvas.drawString(1.5 * cm, 1.1 * cm, f"DocPresupuestoAI · {titulo_doc}")
    canvas.drawRightString(
        A4[0] - 1.5 * cm,
        1.1 * cm,
        f"Página {canvas.getPageNumber()} · {datetime.now().strftime('%d/%m/%Y %H:%M')}",
    )
    canvas.restoreState()


def _clean_inline_markdown(texto: str) -> str:
    line = (texto or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    while "**" in line:
        line = line.replace("**", "<b>", 1)
        if "**" in line:
            line = line.replace("**", "</b>", 1)
    return line


def _seccion_firma_table() -> Table:
    data = [
        ["Aprobación Comercial", "", "Aprobación Técnica", ""],
        ["Nombre:", "__________________________", "Nombre:", "__________________________"],
        ["Cargo:", "__________________________", "Cargo:", "__________________________"],
        ["Firma:", "__________________________", "Firma:", "__________________________"],
    ]
    t = Table(data, colWidths=[3.6 * cm, 5.2 * cm, 3.6 * cm, 5.2 * cm])
    t.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("TEXTCOLOR", (0, 0), (-1, 0), COLOR_PRIMARIO),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("GRID", (0, 1), (-1, -1), 0.25, colors.lightgrey),
    ]))
    return t


def format_clp(value: float) -> str:
    return f"$ {value:,.0f}".replace(",", ".")

def format_uf(value: float) -> str:
    return f"UF {value:,.2f}"

# ═══════════════════════════════════════════════════════════════════════════════
# GENERADOR DE PRESUPUESTO PDF
# ═══════════════════════════════════════════════════════════════════════════════
def generar_presupuesto_pdf(datos: dict, output_path: str) -> str:
    doc = SimpleDocTemplate(
        output_path, pagesize=A4,
        rightMargin=1.5*cm, leftMargin=1.5*cm,
        topMargin=2*cm, bottomMargin=2*cm
    )
    
    styles = getSampleStyleSheet()
    
    # Estilos personalizados (enfoque ejecutivo/profesional)
    estilo_titulo = ParagraphStyle('Titulo', fontSize=20, textColor=COLOR_PRIMARIO,
                                    alignment=TA_CENTER, spaceAfter=6, fontName='Helvetica-Bold')
    estilo_subtitulo = ParagraphStyle('Subtitulo', fontSize=13, textColor=COLOR_SECUNDARIO,
                                       alignment=TA_CENTER, spaceAfter=4, fontName='Helvetica')
    estilo_seccion = ParagraphStyle('Seccion', fontSize=11, textColor=colors.white,
                                     spaceAfter=4, fontName='Helvetica-Bold', leftIndent=8)
    estilo_normal = ParagraphStyle('Normal2', fontSize=9, spaceAfter=2,
                                    fontName='Helvetica', textColor=COLOR_TEXTO)
    estilo_footer = ParagraphStyle('Footer', fontSize=7, textColor=colors.grey,
                                    alignment=TA_CENTER)
    
    story = []
    resumen = datos.get("resumen", {})
    partidas = datos.get("partidas", [])
    
    # ─── PORTADA EJECUTIVA ────────────────────────────────────────────────────
    nombre_proyecto = resumen.get("nombre_proyecto", "Proyecto")
    cliente = resumen.get("cliente", "Cliente no informado")
    fecha_str = resumen.get("fecha", datetime.now().strftime("%d/%m/%Y"))
    moneda = resumen.get("moneda", "CLP")

    portada = Table([["PROPUESTA ECONÓMICA"]], colWidths=[18 * cm])
    portada.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), COLOR_PRIMARIO),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 18),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, -1), 16),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 16),
    ]))
    story.append(portada)
    story.append(Spacer(1, 0.35 * cm))
    story.append(Paragraph(f"<b>{nombre_proyecto}</b>", estilo_titulo))
    story.append(Paragraph(f"Cliente / Mandante: <b>{cliente}</b>", ParagraphStyle("PortadaMeta", fontSize=10, alignment=TA_CENTER)))
    story.append(Paragraph(f"Fecha de emisión: <b>{fecha_str}</b>", ParagraphStyle("PortadaMeta2", fontSize=9, alignment=TA_CENTER, textColor=colors.grey)))
    story.append(Spacer(1, 0.5 * cm))

    # ─── ENCABEZADO DETALLE ───────────────────────────────────────────────────
    banner_data = [["PRESUPUESTO TÉCNICO - OFERTA ECONÓMICA"]]
    banner = Table(banner_data, colWidths=[18*cm])
    banner.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), COLOR_PRIMARIO),
        ('TEXTCOLOR', (0,0), (-1,-1), colors.white),
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 16),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 15),
        ('BOTTOMPADDING', (0,0), (-1,-1), 15),
        ('ROWBACKGROUNDS', (0,0), (-1,-1), [COLOR_PRIMARIO]),
    ]))
    story.append(banner)
    story.append(Spacer(1, 0.3*cm))
    
    # Datos del proyecto
    info_data = [
        ["Proyecto:", nombre_proyecto, "Fecha:", fecha_str],
        ["Cliente:", cliente, "Moneda:", moneda],
    ]
    info_table = Table(info_data, colWidths=[3*cm, 8*cm, 2.5*cm, 4.5*cm])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
        ('FONTNAME', (2,0), (2,-1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0,0), (0,-1), COLOR_PRIMARIO),
        ('TEXTCOLOR', (2,0), (2,-1), COLOR_PRIMARIO),
        ('BACKGROUND', (0,0), (-1,-1), COLOR_FONDO),
        ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('LEFTPADDING', (0,0), (-1,-1), 8),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 0.5*cm))

    # Declaración ejecutiva de alcance y validez
    story.append(Paragraph(
        "Alcance ejecutivo: esta propuesta económica considera la información "
        "disponible al momento de emisión y se alinea al estándar técnico-comercial "
        "definido para procesos de licitación.",
        ParagraphStyle("AlcanceExec", fontSize=8.7, leading=12, textColor=COLOR_TEXTO, alignment=TA_JUSTIFY, spaceAfter=8),
    ))
    
    # ─── TABLA DE PARTIDAS ────────────────────────────────────────────────────
    # Encabezado de tabla
    header_data = [["N°", "PARTIDA / DESCRIPCIÓN", "UNIDAD", "CANTIDAD", "P. UNITARIO", "P. TOTAL"]]
    col_widths = [1*cm, 7.5*cm, 1.8*cm, 1.8*cm, 2.8*cm, 3.1*cm]
    
    table_data = header_data.copy()
    
    categoria_actual = ""
    for idx, item in enumerate(partidas):
        cat = item.get("categoria", "")
        if cat and cat != categoria_actual:
            categoria_actual = cat
            cat_row = [f"  ▸ {cat.upper()}", "", "", "", "", ""]
            table_data.append(cat_row)
        
        pu = item.get("precio_unitario", 0) or 0
        qty = item.get("cantidad", 0) or 0
        pt = item.get("precio_total", 0) or (qty * pu)
        
        row = [
            str(item.get("numero", idx+1)),
            Paragraph(f"<b>{item.get('partida','')}</b><br/><font size='7' color='grey'>{item.get('descripcion','')[:120]}</font>",
                      ParagraphStyle('cell', fontSize=8, leading=11)),
            str(item.get("unidad", "gl")),
            f"{qty:,.2f}",
            format_clp(pu) if moneda == "CLP" else format_uf(pu),
            format_clp(pt) if moneda == "CLP" else format_uf(pt),
        ]
        table_data.append(row)
    
    main_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    table_style = [
        # Encabezado
        ('BACKGROUND', (0,0), (-1,0), COLOR_PRIMARIO),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 9),
        ('ALIGN', (0,0), (-1,0), 'CENTER'),
        ('TOPPADDING', (0,0), (-1,0), 8),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        # Filas alternas
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, COLOR_FONDO]),
        ('FONTSIZE', (0,1), (-1,-1), 8),
        ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN', (2,1), (-1,-1), 'RIGHT'),
        ('ALIGN', (0,1), (0,-1), 'CENTER'),
        ('GRID', (0,0), (-1,-1), 0.3, colors.lightgrey),
        ('TOPPADDING', (0,1), (-1,-1), 4),
        ('BOTTOMPADDING', (0,1), (-1,-1), 4),
        ('LEFTPADDING', (0,0), (-1,-1), 5),
        ('RIGHTPADDING', (0,0), (-1,-1), 5),
    ]
    
    # Resaltar filas de categoría
    for i, row in enumerate(table_data):
        if i > 0 and isinstance(row[0], str) and row[0].startswith("  ▸"):
            table_style.append(('BACKGROUND', (0,i), (-1,i), COLOR_SECUNDARIO))
            table_style.append(('TEXTCOLOR', (0,i), (-1,i), colors.white))
            table_style.append(('FONTNAME', (0,i), (-1,i), 'Helvetica-Bold'))
            table_style.append(('SPAN', (0,i), (-1,i)))
    
    main_table.setStyle(TableStyle(table_style))
    story.append(main_table)
    story.append(Spacer(1, 0.5*cm))
    
    # ─── RESUMEN FINANCIERO ───────────────────────────────────────────────────
    subtotal = resumen.get("subtotal", sum(p.get("precio_total", 0) or 0 for p in partidas))
    gastos_gen = resumen.get("gastos_generales", subtotal * 0.15)
    utilidades = resumen.get("utilidades", subtotal * 0.10)
    neto = resumen.get("neto", subtotal + gastos_gen + utilidades)
    iva = resumen.get("iva", neto * 0.19)
    total = resumen.get("total", neto + iva)
    
    fmt = format_clp if moneda == "CLP" else format_uf
    
    resumen_data = [
        ["", "SUBTOTAL DIRECTO:", fmt(subtotal)],
        ["", "GASTOS GENERALES (15%):", fmt(gastos_gen)],
        ["", "UTILIDADES (10%):", fmt(utilidades)],
        ["", "VALOR NETO:", fmt(neto)],
        ["", "IVA (19%):", fmt(iva)],
        ["", "TOTAL OFERTA:", fmt(total)],
    ]
    
    resumen_table = Table(resumen_data, colWidths=[9.5*cm, 5*cm, 3.5*cm])
    resumen_table.setStyle(TableStyle([
        ('FONTNAME', (1,0), (1,-1), 'Helvetica-Bold'),
        ('FONTNAME', (2,0), (2,-1), 'Helvetica'),
        ('ALIGN', (1,0), (-1,-1), 'RIGHT'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        # Total destacado
        ('BACKGROUND', (1,5), (-1,5), COLOR_ACENTO),
        ('FONTNAME', (1,5), (-1,5), 'Helvetica-Bold'),
        ('FONTSIZE', (1,5), (-1,5), 11),
        ('TEXTCOLOR', (1,5), (-1,5), COLOR_PRIMARIO),
        ('LINEABOVE', (1,3), (-1,3), 1, COLOR_PRIMARIO),
        ('LINEABOVE', (1,5), (-1,5), 1.5, COLOR_PRIMARIO),
    ]))
    story.append(resumen_table)
    
    # ─── NOTAS, VIGENCIA Y FIRMA ─────────────────────────────────────────────
    story.append(Spacer(1, 0.45 * cm))
    vigencia_data = [
        ["Condiciones comerciales", "Detalle"],
        ["Validez de la oferta", resumen.get("validez_oferta", "30 días corridos desde la fecha de emisión")],
        ["Plazo de ejecución", resumen.get("plazo_ejecucion", "Según programación contractual acordada")],
        ["Supuestos relevantes", resumen.get("supuestos", "Sujeto a confirmación de alcance, accesos, y disponibilidad operativa del cliente")],
    ]
    vigencia_table = Table(vigencia_data, colWidths=[4.8 * cm, 13.2 * cm])
    vigencia_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), COLOR_SECUNDARIO),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(vigencia_table)
    story.append(Spacer(1, 0.45 * cm))
    story.append(_seccion_firma_table())

    # ─── FOOTER ───────────────────────────────────────────────────────────────
    story.append(Spacer(1, 0.4*cm))
    story.append(HRFlowable(width="100%", thickness=1, color=COLOR_PRIMARIO))
    story.append(Spacer(1, 0.2*cm))
    story.append(Paragraph(
        f"Documento generado automáticamente por DocPresupuestoAI | {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        estilo_footer
    ))
    
    doc.build(
        story,
        onFirstPage=lambda c, d: _draw_header_footer(c, d, "Propuesta Económica"),
        onLaterPages=lambda c, d: _draw_header_footer(c, d, "Propuesta Económica"),
    )
    return output_path


# ═══════════════════════════════════════════════════════════════════════════════
# GENERADOR EXCEL
# ═══════════════════════════════════════════════════════════════════════════════
def generar_presupuesto_excel(datos: dict, output_path: str) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Presupuesto"
    ws_exec = wb.create_sheet("Resumen Ejecutivo", 0)
    
    resumen = datos.get("resumen", {})
    partidas = datos.get("partidas", [])
    nombre_proyecto = resumen.get("nombre_proyecto", "Proyecto")
    cliente = resumen.get("cliente", "Cliente no informado")
    fecha_emision = resumen.get("fecha", datetime.now().strftime("%d/%m/%Y"))
    moneda = resumen.get("moneda", "CLP")
    
    # Estilos
    azul_oscuro = "1a3c5e"
    azul_medio = "2e7d9b"
    dorado = "f5a623"
    gris_claro = "f0f4f8"
    
    header_font = Font(name='Calibri', bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(fill_type="solid", fgColor=azul_oscuro)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    border_thin = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    
    # ─── Hoja ejecutiva (nueva) ──────────────────────────────────────────────
    ws_exec.merge_cells("A1:E1")
    ws_exec["A1"] = "RESUMEN EJECUTIVO - PROPUESTA ECONÓMICA"
    ws_exec["A1"].font = Font(name='Calibri', bold=True, color="FFFFFF", size=15)
    ws_exec["A1"].fill = PatternFill(fill_type="solid", fgColor=azul_oscuro)
    ws_exec["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_exec.row_dimensions[1].height = 34

    ws_exec["A3"] = "Proyecto:"
    ws_exec["B3"] = nombre_proyecto
    ws_exec["A4"] = "Cliente:"
    ws_exec["B4"] = cliente
    ws_exec["A5"] = "Fecha emisión:"
    ws_exec["B5"] = fecha_emision
    ws_exec["A6"] = "Moneda:"
    ws_exec["B6"] = moneda
    for c in ["A3", "A4", "A5", "A6"]:
        ws_exec[c].font = Font(name='Calibri', bold=True, color=azul_oscuro, size=10)

    # ─── Fila de título (detalle presupuesto) ────────────────────────────────
    ws.merge_cells("A1:F1")
    ws["A1"] = f"PRESUPUESTO: {nombre_proyecto.upper()}"
    ws["A1"].font = Font(name='Calibri', bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor=azul_oscuro)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35
    
    # Datos del proyecto
    ws["A2"] = "Fecha:"
    ws["B2"] = datetime.now().strftime("%d/%m/%Y")
    ws["C2"] = "Moneda:"
    ws["D2"] = moneda
    ws["A2"].font = ws["C2"].font = Font(bold=True, color=azul_oscuro)
    
    ws.row_dimensions[3].height = 5  # Espacio
    
    # ─── Encabezados de tabla ───
    headers = ["N°", "PARTIDA", "DESCRIPCIÓN", "UNIDAD", "CANTIDAD", "P. UNITARIO", "P. TOTAL"]
    col_widths = [5, 25, 40, 10, 12, 18, 18]
    
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border_thin
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[4].height = 22
    
    # ─── Filas de datos ───
    row_num = 5
    for idx, item in enumerate(partidas):
        pu = item.get("precio_unitario", 0) or 0
        qty = item.get("cantidad", 0) or 0
        pt = item.get("precio_total", 0) or (qty * pu)
        
        fill_color = gris_claro if idx % 2 == 0 else "FFFFFF"
        row_fill = PatternFill(fill_type="solid", fgColor=fill_color)
        
        values = [
            str(item.get("numero", idx+1)),
            item.get("partida", ""),
            item.get("descripcion", ""),
            item.get("unidad", "gl"),
            qty,
            pu,
            pt
        ]
        
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.font = Font(name='Calibri', size=9)
            cell.fill = row_fill
            cell.border = border_thin
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col in [5, 6, 7]:
                cell.number_format = '#,##0' if moneda == "CLP" else '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
        
        ws.row_dimensions[row_num].height = 20
        row_num += 1
    
    # ─── Resumen financiero ───
    row_num += 1
    subtotal = sum(p.get("precio_total", 0) or 0 for p in partidas)
    gastos_gen = resumen.get("gastos_generales", subtotal * 0.15)
    utilidades = resumen.get("utilidades", subtotal * 0.10)
    neto = resumen.get("neto", subtotal + gastos_gen + utilidades)
    iva = resumen.get("iva", neto * 0.19)
    total = resumen.get("total", neto + iva)
    
    resumen_rows = [
        ("Subtotal Directo:", subtotal),
        ("Gastos Generales (15%):", gastos_gen),
        ("Utilidades (10%):", utilidades),
        ("Valor Neto:", neto),
        ("IVA (19%):", iva),
        ("TOTAL OFERTA:", total),
    ]
    
    for label, value in resumen_rows:
        ws.merge_cells(f"A{row_num}:E{row_num}")
        label_cell = ws.cell(row=row_num, column=1, value=label)
        value_cell = ws.cell(row=row_num, column=6, value=value)
        label_cell.font = Font(bold=True, color=azul_oscuro)
        label_cell.alignment = Alignment(horizontal="right")
        value_cell.number_format = '#,##0' if moneda == "CLP" else '#,##0.00'
        value_cell.alignment = Alignment(horizontal="right")
        
        if label == "TOTAL OFERTA:":
            for col in range(1, 8):
                ws.cell(row=row_num, column=col).fill = PatternFill(fill_type="solid", fgColor=dorado)
                ws.cell(row=row_num, column=col).font = Font(bold=True, size=11, color=azul_oscuro)
        
        row_num += 1
    
    # Congelar paneles
    ws.freeze_panes = "A5"

    # ─── KPI ejecutivos y condiciones comerciales ────────────────────────────
    ws_exec["A8"] = "Indicador"
    ws_exec["B8"] = "Valor"
    ws_exec["C8"] = "Referencia"
    for col in ["A", "B", "C"]:
        cell = ws_exec[f"{col}8"]
        cell.font = Font(name='Calibri', bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill(fill_type="solid", fgColor=azul_oscuro)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    indicadores = [
        ("Subtotal directo", subtotal, "Base técnica"),
        ("Gastos generales", gastos_gen, "Administración y soporte"),
        ("Utilidades", utilidades, "Margen objetivo"),
        ("Valor neto", neto, "Sin IVA"),
        ("IVA", iva, "19%"),
        ("Total oferta", total, "Monto final"),
    ]
    fila_exec = 9
    for idx, (nombre, valor, ref) in enumerate(indicadores):
        ws_exec[f"A{fila_exec}"] = nombre
        ws_exec[f"B{fila_exec}"] = valor
        ws_exec[f"C{fila_exec}"] = ref
        ws_exec[f"B{fila_exec}"].number_format = '#,##0' if moneda == "CLP" else '#,##0.00'
        ws_exec[f"A{fila_exec}"].font = Font(name='Calibri', bold=(nombre == "Total oferta"), color=azul_oscuro)
        if nombre == "Total oferta":
            for c in ["A", "B", "C"]:
                ws_exec[f"{c}{fila_exec}"].fill = PatternFill(fill_type="solid", fgColor=dorado)
                ws_exec[f"{c}{fila_exec}"].font = Font(name='Calibri', bold=True, color=azul_oscuro, size=11)
        else:
            fill = gris_claro if idx % 2 == 0 else "FFFFFF"
            for c in ["A", "B", "C"]:
                ws_exec[f"{c}{fila_exec}"].fill = PatternFill(fill_type="solid", fgColor=fill)
        for c in ["A", "B", "C"]:
            ws_exec[f"{c}{fila_exec}"].alignment = Alignment(horizontal="left" if c != "B" else "right", vertical="center")
            ws_exec[f"{c}{fila_exec}"].border = border_thin
        fila_exec += 1

    fila_exec += 1
    ws_exec.merge_cells(f"A{fila_exec}:C{fila_exec}")
    ws_exec[f"A{fila_exec}"] = "Condiciones comerciales"
    ws_exec[f"A{fila_exec}"].font = Font(name='Calibri', bold=True, color="FFFFFF", size=10)
    ws_exec[f"A{fila_exec}"].fill = PatternFill(fill_type="solid", fgColor=azul_medio)
    ws_exec[f"A{fila_exec}"].alignment = Alignment(horizontal="left", vertical="center")

    condiciones = [
        f"Validez de oferta: {resumen.get('validez_oferta', '30 días corridos desde emisión')}",
        f"Plazo de ejecución: {resumen.get('plazo_ejecucion', 'Según programación contractual')}",
        f"Supuestos relevantes: {resumen.get('supuestos', 'Sujeto a validación de alcance, accesos y ventanas operativas')}",
    ]
    for item in condiciones:
        fila_exec += 1
        ws_exec.merge_cells(f"A{fila_exec}:C{fila_exec}")
        ws_exec[f"A{fila_exec}"] = f"• {item}"
        ws_exec[f"A{fila_exec}"].font = Font(name='Calibri', size=9, color="1a1a2e")
        ws_exec[f"A{fila_exec}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws_exec[f"A{fila_exec}"].fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
        ws_exec[f"A{fila_exec}"].border = border_thin

    ws_exec.column_dimensions["A"].width = 34
    ws_exec.column_dimensions["B"].width = 20
    ws_exec.column_dimensions["C"].width = 34
    ws_exec.freeze_panes = "A9"
    
    wb.save(output_path)
    return output_path


# ═══════════════════════════════════════════════════════════════════════════════
# GENERADOR DE INFORME TÉCNICO PDF
# ═══════════════════════════════════════════════════════════════════════════════
def generar_informe_pdf(contenido_md: str, datos_proyecto: dict, output_path: str) -> str:
    doc = SimpleDocTemplate(
        output_path, pagesize=A4,
        rightMargin=2*cm, leftMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm
    )
    
    styles = getSampleStyleSheet()
    
    estilo_h1 = ParagraphStyle('H1', fontSize=18, textColor=COLOR_PRIMARIO,
                                spaceAfter=12, fontName='Helvetica-Bold',
                                alignment=TA_CENTER)
    estilo_h2 = ParagraphStyle('H2', fontSize=13, textColor=COLOR_SECUNDARIO,
                                spaceAfter=8, spaceBefore=12, fontName='Helvetica-Bold',
                                borderPadding=(5,0,5,0))
    estilo_h3 = ParagraphStyle('H3', fontSize=11, textColor=COLOR_PRIMARIO,
                                spaceAfter=6, spaceBefore=8, fontName='Helvetica-Bold')
    estilo_cuerpo = ParagraphStyle('Cuerpo', fontSize=10, leading=15,
                                    spaceAfter=8, fontName='Helvetica',
                                    alignment=TA_JUSTIFY)
    estilo_bullet = ParagraphStyle('Bullet', fontSize=10, leading=14,
                                    leftIndent=20, spaceAfter=4,
                                    bulletIndent=10, fontName='Helvetica')
    
    story = []
    
    # Portada ejecutiva
    story.append(Spacer(1, 2*cm))
    
    banner = Table([["INFORME TÉCNICO"]], colWidths=[17*cm])
    banner.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), COLOR_PRIMARIO),
        ('TEXTCOLOR', (0,0), (-1,-1), colors.white),
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 22),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('TOPPADDING', (0,0), (-1,-1), 20),
        ('BOTTOMPADDING', (0,0), (-1,-1), 20),
    ]))
    story.append(banner)
    story.append(Spacer(1, 0.5*cm))
    
    proyecto_obj = datos_proyecto.get("proyecto", {}) or {}
    proyecto_nombre = proyecto_obj.get("nombre", "Proyecto")
    cliente = proyecto_obj.get("mandante", "Cliente no informado")
    modalidad = proyecto_obj.get("modalidad", "No informada")
    plazo = proyecto_obj.get("plazo_ejecucion", "No informado")
    story.append(Paragraph(proyecto_nombre, estilo_h1))
    story.append(Paragraph(
        f"Cliente / Mandante: <b>{cliente}</b> · Modalidad: <b>{modalidad}</b> · Plazo: <b>{plazo}</b>",
        ParagraphStyle('meta_info', fontSize=9, alignment=TA_CENTER, textColor=COLOR_TEXTO, spaceAfter=6),
    ))
    story.append(Spacer(1, 0.3*cm))
    story.append(Paragraph(f"Fecha: {_fmt_fecha_larga()}", 
                           ParagraphStyle('fecha', fontSize=10, alignment=TA_CENTER, 
                                          textColor=colors.grey)))
    story.append(Spacer(1, 0.3*cm))
    story.append(Paragraph(
        "Este informe presenta alcance, metodología y criterios técnicos con enfoque "
        "de toma de decisión ejecutiva para procesos de licitación.",
        ParagraphStyle('portada_desc', fontSize=9, alignment=TA_CENTER, textColor=COLOR_TEXTO, leading=13),
    ))
    story.append(PageBreak())
    
    # Contenido Markdown → PDF
    lines = contenido_md.split('\n')
    for line in lines:
        line = line.strip()
        if not line:
            story.append(Spacer(1, 0.3*cm))
        elif line.startswith('# '):
            story.append(Paragraph(line[2:], estilo_h1))
        elif line.startswith('## '):
            # Sección con fondo de color
            sect = Table([[line[3:]]], colWidths=[17*cm])
            sect.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,-1), COLOR_SECUNDARIO),
                ('TEXTCOLOR', (0,0), (-1,-1), colors.white),
                ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,-1), 12),
                ('LEFTPADDING', (0,0), (-1,-1), 10),
                ('TOPPADDING', (0,0), (-1,-1), 6),
                ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ]))
            story.append(Spacer(1, 0.3*cm))
            story.append(sect)
            story.append(Spacer(1, 0.2*cm))
        elif line.startswith('### '):
            story.append(Paragraph(line[4:], estilo_h3))
        elif line.startswith('- ') or line.startswith('* '):
            story.append(Paragraph(f"• {_clean_inline_markdown(line[2:])}", estilo_bullet))
        elif line.startswith('**') and line.endswith('**'):
            story.append(Paragraph(f"<b>{_clean_inline_markdown(line[2:-2])}</b>", estilo_cuerpo))
        else:
            story.append(Paragraph(_clean_inline_markdown(line), estilo_cuerpo))

    # Cierre formal con supuestos y firmas
    story.append(Spacer(1, 0.5 * cm))
    clausulas = [
        ["Declaración", "Contenido"],
        ["Alcance", "La presente propuesta técnica se basa en antecedentes disponibles y puede ajustarse por cambios de alcance."],
        ["Riesgos y restricciones", "Sujeto a accesos a terreno, ventanas operativas y validación documental de cliente."],
        ["Control de cambios", "Toda modificación relevante debe ser registrada y aprobada por ambas partes."],
    ]
    clausulas_table = Table(clausulas, colWidths=[4.6 * cm, 12.4 * cm])
    clausulas_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), COLOR_SECUNDARIO),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(clausulas_table)
    story.append(Spacer(1, 0.45 * cm))
    story.append(_seccion_firma_table())

    story.append(Spacer(1, 1*cm))
    story.append(HRFlowable(width="100%", thickness=1, color=COLOR_PRIMARIO))
    story.append(Paragraph(
        f"Generado por DocPresupuestoAI | {datetime.now().strftime('%d/%m/%Y')}",
        ParagraphStyle('footer', fontSize=7, textColor=colors.grey, alignment=TA_CENTER)
    ))
    
    doc.build(
        story,
        onFirstPage=lambda c, d: _draw_header_footer(c, d, "Informe Técnico"),
        onLaterPages=lambda c, d: _draw_header_footer(c, d, "Informe Técnico"),
    )
    return output_path


# ═══════════════════════════════════════════════════════════════════════════════
# INDICE DOCUMENTAL (PDF / EXCEL)
# ═══════════════════════════════════════════════════════════════════════════════
def generar_indice_documental_excel(
    proyecto_nombre: str,
    secciones_ordenadas: list,
    output_path: str,
    cliente: str = "",
    fecha_generacion: str = "",
    codigo_licitacion: str = ""
) -> str:
    wb = openpyxl.Workbook()
    ws_resumen = wb.active
    ws_resumen.title = "Resumen Dossier"
    ws = wb.create_sheet("Indice Documental")

    azul_oscuro = "1a3c5e"
    gris_claro = "f0f4f8"
    verde = "27ae60"
    dorado = "f5a623"

    if not fecha_generacion:
        fecha_generacion = datetime.now().strftime("%d/%m/%Y %H:%M")

    # ─── Hoja Resumen Dossier ────────────────────────────────────────────────
    ws_resumen.merge_cells("A1:E1")
    ws_resumen["A1"] = "DOSSIER DOCUMENTAL DE LICITACION"
    ws_resumen["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=15)
    ws_resumen["A1"].fill = PatternFill(fill_type="solid", fgColor=azul_oscuro)
    ws_resumen["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_resumen.row_dimensions[1].height = 34

    ws_resumen["A3"] = "Proyecto:"
    ws_resumen["B3"] = proyecto_nombre
    ws_resumen["A4"] = "Cliente:"
    ws_resumen["B4"] = cliente or "No informado"
    ws_resumen["A5"] = "Codigo OT/Licitacion:"
    ws_resumen["B5"] = codigo_licitacion or "No informado"
    ws_resumen["A6"] = "Fecha de emisión:"
    ws_resumen["B6"] = fecha_generacion
    for c in ["A3", "A4", "A5", "A6"]:
        ws_resumen[c].font = Font(name="Calibri", bold=True, color=azul_oscuro, size=10)

    ws_resumen["A8"] = "Sección"
    ws_resumen["B8"] = "Total requisitos"
    ws_resumen["C8"] = "Cumplidos"
    ws_resumen["D8"] = "% Cumplimiento"
    ws_resumen["E8"] = "Estado"
    for col in ["A", "B", "C", "D", "E"]:
        cell = ws_resumen[f"{col}8"]
        cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill(fill_type="solid", fgColor=azul_oscuro)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    fila = 9
    total_global = 0
    cumplidos_global = 0
    for sec in secciones_ordenadas:
        reqs = sec.get("requisitos", []) or []
        total_sec = len(reqs)
        cumplidos_sec = sum(1 for r in reqs if (r.get("estado") or "").strip().lower() == "cumplido")
        pct = (cumplidos_sec / total_sec * 100) if total_sec else 0
        estado = "OK" if pct >= 90 else ("PARCIAL" if pct >= 60 else "CRITICO")

        ws_resumen[f"A{fila}"] = sec.get("seccion", "")
        ws_resumen[f"B{fila}"] = total_sec
        ws_resumen[f"C{fila}"] = cumplidos_sec
        ws_resumen[f"D{fila}"] = round(pct, 2)
        ws_resumen[f"E{fila}"] = estado
        ws_resumen[f"D{fila}"].number_format = '0.00"%"'
        ws_resumen[f"A{fila}"].font = Font(name="Calibri", bold=True, color=azul_oscuro)

        fill = gris_claro if fila % 2 == 0 else "FFFFFF"
        for col in ["A", "B", "C", "D", "E"]:
            ws_resumen[f"{col}{fila}"].fill = PatternFill(fill_type="solid", fgColor=fill)
            ws_resumen[f"{col}{fila}"].alignment = Alignment(horizontal="center", vertical="center")

        if estado == "OK":
            ws_resumen[f"E{fila}"].fill = PatternFill(fill_type="solid", fgColor=verde)
            ws_resumen[f"E{fila}"].font = Font(name="Calibri", bold=True, color="FFFFFF")
        elif estado == "PARCIAL":
            ws_resumen[f"E{fila}"].fill = PatternFill(fill_type="solid", fgColor=dorado)
            ws_resumen[f"E{fila}"].font = Font(name="Calibri", bold=True, color=azul_oscuro)

        total_global += total_sec
        cumplidos_global += cumplidos_sec
        fila += 1

    pct_global = (cumplidos_global / total_global * 100) if total_global else 0
    ws_resumen[f"A{fila+1}"] = "Resumen Global"
    ws_resumen[f"B{fila+1}"] = total_global
    ws_resumen[f"C{fila+1}"] = cumplidos_global
    ws_resumen[f"D{fila+1}"] = round(pct_global, 2)
    ws_resumen[f"D{fila+1}"].number_format = '0.00"%"'
    for col in ["A", "B", "C", "D"]:
        ws_resumen[f"{col}{fila+1}"].font = Font(name="Calibri", bold=True, color=azul_oscuro, size=11)

    ws_resumen.column_dimensions["A"].width = 20
    ws_resumen.column_dimensions["B"].width = 18
    ws_resumen.column_dimensions["C"].width = 14
    ws_resumen.column_dimensions["D"].width = 16
    ws_resumen.column_dimensions["E"].width = 14

    ws.merge_cells("A1:G1")
    ws["A1"] = f"INDICE DOCUMENTAL - {proyecto_nombre}"
    ws["A1"].font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor=azul_oscuro)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    headers = ["Seccion", "Requisito", "Categoria", "Estado", "Carpeta", "Nro Evidencias", "Archivos"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill(fill_type="solid", fgColor=azul_oscuro)
        c.alignment = Alignment(horizontal="center", vertical="center")

    row = 4
    for sec in secciones_ordenadas:
        seccion = sec.get("seccion", "")
        requisitos = sec.get("requisitos", []) or []
        if not requisitos:
            ws.cell(row=row, column=1, value=seccion)
            ws.cell(row=row, column=2, value="(Sin requisitos)")
            row += 1
            continue

        for req in requisitos:
            evidencias = req.get("evidencias", []) or []
            ws.cell(row=row, column=1, value=seccion)
            ws.cell(row=row, column=2, value=req.get("requisito", ""))
            ws.cell(row=row, column=3, value=req.get("categoria", ""))
            ws.cell(row=row, column=4, value=req.get("estado", ""))
            ws.cell(row=row, column=5, value=req.get("carpeta_objetivo", ""))
            ws.cell(row=row, column=6, value=len(evidencias))
            ws.cell(row=row, column=7, value=", ".join(e.get("nombre_archivo", "") for e in evidencias))
            row += 1

    for r in range(4, row):
        if r % 2 == 0:
            for c in range(1, 8):
                ws.cell(row=r, column=c).fill = PatternFill(fill_type="solid", fgColor=gris_claro)
        for c in range(1, 8):
            ws.cell(row=r, column=c).alignment = Alignment(vertical="top", wrap_text=True)
            ws.cell(row=r, column=c).font = Font(name="Calibri", size=9)

    widths = [16, 34, 16, 14, 40, 14, 45]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"
    wb.save(output_path)
    return output_path


def generar_indice_documental_pdf(
    proyecto_nombre: str,
    secciones_ordenadas: list,
    output_path: str,
    cliente: str = "",
    fecha_generacion: str = "",
    logo_path: str = "",
    codigo_licitacion: str = ""
) -> str:
    doc = SimpleDocTemplate(
        output_path,
        pagesize=landscape(A4),
        rightMargin=1.2 * cm,
        leftMargin=1.2 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )

    styles = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle(
        "TituloIndice",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=15,
        alignment=TA_CENTER,
        textColor=COLOR_PRIMARIO,
        spaceAfter=8,
    )
    estilo_normal = ParagraphStyle(
        "NormalIndice",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=8,
        leading=10,
    )

    story = []
    if not fecha_generacion:
        fecha_generacion = datetime.now().strftime("%d/%m/%Y %H:%M")

    # ─── Portada tipo dossier ────────────────────────────────────────────────
    portada = Table([["DOSSIER DOCUMENTAL DE LICITACION"]], colWidths=[26.5 * cm])
    portada.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), COLOR_PRIMARIO),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 18),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 16),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 16),
    ]))
    story.append(portada)
    story.append(Spacer(1, 0.5 * cm))
    if logo_path and os.path.exists(logo_path):
        try:
            logo = Image(logo_path, width=4.2 * cm, height=1.6 * cm)
            logo.hAlign = "CENTER"
            story.append(logo)
            story.append(Spacer(1, 0.25 * cm))
        except Exception:
            pass
    story.append(Paragraph(f"Proyecto: <b>{proyecto_nombre}</b>", ParagraphStyle("meta1", parent=estilo_normal, fontSize=10)))
    story.append(Paragraph(f"Cliente: <b>{cliente or 'No informado'}</b>", ParagraphStyle("meta2", parent=estilo_normal, fontSize=10)))
    story.append(Paragraph(f"Codigo OT/Licitacion: <b>{codigo_licitacion or 'No informado'}</b>", ParagraphStyle("meta3", parent=estilo_normal, fontSize=10)))
    story.append(Paragraph(f"Fecha de emisión: <b>{fecha_generacion}</b>", ParagraphStyle("meta4", parent=estilo_normal, fontSize=10)))
    story.append(Spacer(1, 0.5 * cm))

    # ─── Resumen ejecutivo por seccion ───────────────────────────────────────
    resumen_data = [["Seccion", "Total requisitos", "Cumplidos", "% Cumplimiento", "Estado"]]
    total_global = 0
    cumplidos_global = 0
    for sec in secciones_ordenadas:
        reqs = sec.get("requisitos", []) or []
        total_sec = len(reqs)
        cumplidos_sec = sum(1 for r in reqs if (r.get("estado") or "").strip().lower() == "cumplido")
        pct = (cumplidos_sec / total_sec * 100) if total_sec else 0
        estado = "OK" if pct >= 90 else ("PARCIAL" if pct >= 60 else "CRITICO")
        resumen_data.append([sec.get("seccion", ""), str(total_sec), str(cumplidos_sec), f"{pct:.2f}%", estado])
        total_global += total_sec
        cumplidos_global += cumplidos_sec

    pct_global = (cumplidos_global / total_global * 100) if total_global else 0
    resumen_data.append(["GLOBAL", str(total_global), str(cumplidos_global), f"{pct_global:.2f}%", ""])

    resumen_tabla = Table(resumen_data, colWidths=[5.2 * cm, 4.0 * cm, 3.2 * cm, 4.2 * cm, 3.4 * cm], repeatRows=1)
    resumen_tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), COLOR_PRIMARIO),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, COLOR_FONDO]),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ALIGN", (1, 1), (-1, -1), "CENTER"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(Paragraph("Resumen Ejecutivo", ParagraphStyle("res_h", parent=estilo_titulo, fontSize=12, alignment=TA_LEFT)))
    story.append(resumen_tabla)
    story.append(PageBreak())

    # ─── Detalle del indice ───────────────────────────────────────────────────
    story.append(Paragraph(f"Indice Documental - {proyecto_nombre}", estilo_titulo))
    story.append(Paragraph(f"Generado: {fecha_generacion}", estilo_normal))
    story.append(Spacer(1, 0.3 * cm))

    data = [["Seccion", "Requisito", "Categoria", "Estado", "Carpeta", "Evidencias"]]
    for sec in secciones_ordenadas:
        seccion = sec.get("seccion", "")
        requisitos = sec.get("requisitos", []) or []
        if not requisitos:
            data.append([seccion, "(Sin requisitos)", "", "", "", ""])
            continue
        for req in requisitos:
            evidencias = req.get("evidencias", []) or []
            data.append([
                seccion,
                req.get("requisito", ""),
                req.get("categoria", ""),
                req.get("estado", ""),
                req.get("carpeta_objetivo", ""),
                ", ".join(e.get("nombre_archivo", "") for e in evidencias),
            ])

    tabla = Table(data, colWidths=[3.0 * cm, 5.5 * cm, 2.8 * cm, 2.3 * cm, 8.0 * cm, 6.2 * cm], repeatRows=1)
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), COLOR_PRIMARIO),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, COLOR_FONDO]),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(tabla)
    doc.build(story)
    return output_path


def _xml_para(text) -> str:
    if text is None:
        return ""
    return str(text).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _safe_table_cell(val) -> str:
    """Texto plano para celdas Table (sin entidades XML)."""
    if val is None:
        return ""
    return str(val).replace("\x00", "")


def _fmt_fecha_iso_corta(iso: str | None) -> str:
    if not iso:
        return "—"
    try:
        dt = datetime.fromisoformat(iso.replace("Z", "+00:00"))
        return dt.strftime("%d/%m/%Y")
    except Exception:
        return str(iso)[:10]


def generar_informe_comite_pdf(data: dict, output_path: str) -> str:
    """
    Informe ejecutivo portfolio: KPIs, top riesgos, cartera y plan de cierre (PDF).
    `data` debe alinearse con GET /api/dashboard/ejecutivo (+ semáforos ordenados si aplica).
    """
    doc = SimpleDocTemplate(
        output_path,
        pagesize=A4,
        rightMargin=1.8 * cm,
        leftMargin=1.8 * cm,
        topMargin=1.8 * cm,
        bottomMargin=1.8 * cm,
    )
    estilo_sub = ParagraphStyle(
        "ComiteSub",
        fontSize=10,
        textColor=COLOR_TEXTO,
        alignment=TA_CENTER,
        spaceAfter=14,
        fontName="Helvetica",
    )
    estilo_h2 = ParagraphStyle(
        "ComiteH2",
        fontSize=12,
        textColor=COLOR_SECUNDARIO,
        spaceAfter=8,
        spaceBefore=10,
        fontName="Helvetica-Bold",
    )
    estilo_cuerpo = ParagraphStyle(
        "ComiteCuerpo",
        fontSize=9,
        leading=13,
        spaceAfter=5,
        fontName="Helvetica",
        textColor=COLOR_TEXTO,
    )

    sems = data.get("semaforos") or []
    rojos = sum(1 for x in sems if (x.get("semaforo") or {}).get("color") == "rojo")
    amarillos = sum(1 for x in sems if (x.get("semaforo") or {}).get("color") == "amarillo")
    verdes = sum(1 for x in sems if (x.get("semaforo") or {}).get("color") == "verde")
    aptos = sum(1 for x in sems if (x.get("semaforo") or {}).get("estado_preflight") == "apto")
    total = int(data.get("proyectos_total") or len(sems) or 0)
    aptos_pct = round((aptos / total) * 100, 1) if total else 0.0

    story = []
    banner = Table([["INFORME COMITÉ DE ENTREGA"]], colWidths=[17.4 * cm])
    banner.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), COLOR_PRIMARIO),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 16),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 14),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 14),
    ]))
    story.append(banner)
    story.append(Spacer(1, 0.4 * cm))
    story.append(Paragraph("DocPresupuestoAI · Pulso AI", estilo_sub))
    story.append(
        Paragraph(
            f"Fecha de emisión: {_fmt_fecha_larga()} · Cartera: <b>{total}</b> proyecto(s)",
            estilo_sub,
        )
    )
    story.append(Spacer(1, 0.3 * cm))

    kpi_data = [
        ["Riesgo alto (rojo)", "Atención (amarillo)", "Controlado (verde)", "% Preflight apto"],
        [str(rojos), str(amarillos), str(verdes), f"{aptos_pct}%"],
    ]
    kpi_tab = Table(kpi_data, colWidths=[4.35 * cm, 4.35 * cm, 4.35 * cm, 4.35 * cm])
    kpi_tab.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), COLOR_SECUNDARIO),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 1), (-1, 1), 11),
        ("ROWBACKGROUNDS", (0, 1), (-1, 1), [colors.white, COLOR_FONDO]),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.lightgrey),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(kpi_tab)
    story.append(Spacer(1, 0.5 * cm))

    story.append(Paragraph("Top riesgos operativos (acciones priorizadas)", estilo_h2))
    tops = data.get("top_acciones_criticas") or []
    if not tops:
        story.append(Paragraph("Sin acciones críticas priorizadas en este momento.", estilo_cuerpo))
    else:
        for it in tops:
            tit = _xml_para(it.get("titulo"))
            pn = _xml_para(it.get("proyecto_nombre") or f"Proyecto {it.get('proyecto_id', '')}")
            pr = _xml_para(str(it.get("prioridad") or "").upper())
            ow = _xml_para(it.get("owner") or "equipo")
            fc = _xml_para(_fmt_fecha_iso_corta(it.get("fecha_compromiso")))
            story.append(
                Paragraph(
                    f"• <b>{tit}</b><br/>{pn} · Prioridad {pr} · Owner {ow} · SLA {fc}",
                    estilo_cuerpo,
                )
            )
    story.append(Spacer(1, 0.35 * cm))

    story.append(Paragraph("Cartera de proyectos", estilo_h2))
    rows_p = [["ID", "Proyecto", "OT / Código", "Cliente", "Semáforo", "Preflight", "Venc.", "Avance %"]]
    for s in sems:
        sm = s.get("semaforo") or {}
        rows_p.append([
            str(s.get("proyecto_id", "")),
            _safe_table_cell(s.get("nombre") or ""),
            _safe_table_cell(s.get("codigo_licitacion") or "—"),
            _safe_table_cell(s.get("cliente") or "—"),
            _safe_table_cell(sm.get("label") or "—"),
            _safe_table_cell(str(sm.get("estado_preflight") or "—").upper()),
            str(sm.get("vencidas") if sm.get("vencidas") is not None else "0"),
            f"{float(sm.get('avance_pct') or 0):.0f}%" if sm.get("avance_pct") is not None else "—",
        ])
    if len(rows_p) == 1:
        rows_p.append(["—", "(Sin proyectos)", "", "", "", "", "", ""])
    tab_p = Table(
        rows_p,
        colWidths=[1.1 * cm, 3.6 * cm, 2.6 * cm, 2.5 * cm, 2.2 * cm, 2.0 * cm, 1.0 * cm, 1.3 * cm],
        repeatRows=1,
    )
    tab_p.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), COLOR_PRIMARIO),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 7),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, COLOR_FONDO]),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(tab_p)
    story.append(Spacer(1, 0.45 * cm))

    story.append(Paragraph("Plan de cierre abierto (priorizado)", estilo_h2))
    acc = data.get("acciones_cierre_abiertas") or []
    rows_a = [["Proyecto", "Acción", "Prioridad", "Owner", "SLA", "Estado"]]
    for i in acc:
        est = str(i.get("estado") or "").replace("_", " ")
        rows_a.append([
            _safe_table_cell(i.get("proyecto_nombre") or f"#{i.get('proyecto_id', '')}"),
            _safe_table_cell(i.get("titulo") or ""),
            _safe_table_cell(str(i.get("prioridad") or "").upper()),
            _safe_table_cell(i.get("owner") or "equipo"),
            _safe_table_cell(_fmt_fecha_iso_corta(i.get("fecha_compromiso"))),
            _safe_table_cell(est),
        ])
    if len(rows_a) == 1:
        rows_a.append(["—", "(Sin acciones abiertas)", "", "", "", ""])
    tab_a = Table(
        rows_a,
        colWidths=[3.8 * cm, 5.8 * cm, 2.0 * cm, 2.2 * cm, 2.0 * cm, 2.0 * cm],
        repeatRows=1,
    )
    tab_a.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), COLOR_PRIMARIO),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 7),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, COLOR_FONDO]),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(tab_a)
    story.append(Spacer(1, 0.6 * cm))
    story.append(
        Paragraph(
            "Documento para uso interno. Indicadores según reglas locales de preflight, QA y plan de cierre.",
            ParagraphStyle("ComitePie", fontSize=7, textColor=colors.grey, alignment=TA_CENTER),
        )
    )

    doc.build(
        story,
        onFirstPage=lambda c, d: _draw_header_footer(c, d, "Informe Comité"),
        onLaterPages=lambda c, d: _draw_header_footer(c, d, "Informe Comité"),
    )
    return output_path
