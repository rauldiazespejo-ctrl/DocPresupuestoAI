import os
import sys
import json
import uuid
import shutil
import re
import csv
import io
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

from fastapi import FastAPI, File, UploadFile, Form, HTTPException, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session
import openpyxl

# Agregar paths
sys.path.insert(0, "/Users/rauldiaz/DocPresupuestoAI")

from database.models import (
    create_tables,
    get_db,
    Proyecto,
    Documento,
    ItemPresupuesto,
    OfertaLicitacion,
    PrediccionAdjudicacion,
    RequisitoDocumental,
    EvidenciaRequisito,
    HistoricoLicitacion,
    LegalAceptacionProyecto,
    PlanCierreItem,
)
from backend.extractor import extract_text
from backend.ai_engine import AIEngine
from backend.generator import (
    generar_presupuesto_pdf,
    generar_presupuesto_excel,
    generar_informe_pdf,
    generar_indice_documental_excel,
    generar_indice_documental_pdf,
)
from backend.adjudicacion import calcular_prediccion, calcular_atractividad_licitacion
from backend.ml_atractividad import train_logistic_model, predict_atractividad_ml

# ─── Inicialización ────────────────────────────────────────────────────────────
app = FastAPI(
    title="DocPresupuestoAI",
    description="Sistema inteligente de documentación y presupuestos basado en IA",
    version="1.0.0"
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

BASE_DIR = Path("/Users/rauldiaz/DocPresupuestoAI")
UPLOADS_DIR = BASE_DIR / "uploads"
EXPORTS_DIR = BASE_DIR / "exports"
UPLOADS_DIR.mkdir(exist_ok=True)
EXPORTS_DIR.mkdir(exist_ok=True)

create_tables()

SECCIONES_DOCUMENTALES = [
    ("1_ANT. ECO", "economico"),
    ("2_ANT. ADJ", "adjudicacion"),
    ("3_ANT. TEC", "tecnico"),
    ("4_ANT. PREV", "prevencion"),
    ("5_ANT. ADM", "administrativo"),
    ("6_ANT. FACT", "facturacion"),
]
CODIGO_LICITACION_REGEX = re.compile(r"^OT_[0-9]{8}_[A-Z0-9]+(?:_[A-Z0-9]+)*$")
LEGAL_TERMS_CURRENT_VERSION = "v1.0.0"
PRELIGHT_QA_MIN_SCORE = 85
PREFLIGHT_REQUIRED_DOC_GROUPS = {
    "presupuesto": ["presupuesto_pdf", "presupuesto_excel"],
    "informe": ["informe_pdf"],
    "propuesta": ["propuesta_pdf"],
    "indice_documental": ["indice_documental_pdf", "indice_documental_excel"],
}
PREFLIGHT_BLOCKED_DOWNLOAD_TYPES = {
    "presupuesto_pdf",
    "presupuesto_excel",
    "informe_pdf",
    "propuesta_pdf",
    "indice_documental_pdf",
    "indice_documental_excel",
}


def _slugify(texto: str) -> str:
    limpio = re.sub(r"[^a-zA-Z0-9]+", "_", (texto or "").strip().lower())
    return limpio.strip("_")[:80] or "item"


def _normalizar_codigo_licitacion(codigo: str) -> str:
    return (codigo or "").strip().upper()


def _seccion_documental_por_categoria(categoria: str) -> str:
    cat = (categoria or "").strip().lower()
    if cat in {"economico", "econ", "presupuesto"}:
        return "1_ANT. ECO"
    if cat in {"legal", "adjudicacion", "contrato"}:
        return "2_ANT. ADJ"
    if cat in {"tecnico", "tecnica", "tecnica_operativa"}:
        return "3_ANT. TEC"
    if cat in {"hse", "prevencion", "seguridad"}:
        return "4_ANT. PREV"
    if cat in {"administrativo", "adm"}:
        return "5_ANT. ADM"
    if cat in {"facturacion", "fact", "finanzas"}:
        return "6_ANT. FACT"
    return "5_ANT. ADM"


def _carpeta_requisito(proyecto_id: int, categoria: str, nombre_requisito: str) -> Path:
    seccion = _seccion_documental_por_categoria(categoria)
    return (
        UPLOADS_DIR
        / f"proyecto_{proyecto_id}"
        / "documentacion_requerida"
        / seccion
        / _slugify(nombre_requisito)
    )

# ─── Modelos Pydantic ──────────────────────────────────────────────────────────
class ConfigIA(BaseModel):
    provider: str = "openai"
    api_key: str
    model: str = ""

class GenerarDocumentoRequest(BaseModel):
    proyecto_id: int
    tipo: str  # presupuesto_pdf, presupuesto_excel, informe_pdf, propuesta_pdf
    config_ia: ConfigIA
    quality_mode: str = "standard"  # standard | pro

class ConsultaRequest(BaseModel):
    proyecto_id: int
    pregunta: str
    config_ia: ConfigIA


class OfertaLicitacionRequest(BaseModel):
    proyecto_id: int
    nombre: str
    monto_oferta: float = 0.0
    plazo_dias: int = 0
    factores: dict = Field(default_factory=dict)
    notas: str = ""


class SimulacionEscenario(BaseModel):
    nombre: str
    factores: dict = Field(default_factory=dict)


class SimularPrediccionRequest(BaseModel):
    proyecto_id: int
    escenarios: list[SimulacionEscenario] = Field(default_factory=list)


class BidLevelingRequest(BaseModel):
    proyecto_id: int
    oferta_ids: list[int] = Field(default_factory=list)


class HistoricoLicitacionRequest(BaseModel):
    codigo_licitacion: str = ""
    cliente: str
    rubro: str = ""
    monto_ofertado: float = 0.0
    margen_pct: float = 0.0
    fue_adjudicada: bool = False
    fecha_cierre: str = ""
    observaciones: str = ""


class AtractividadRequest(BaseModel):
    proyecto_id: int
    factores: dict = Field(default_factory=dict)


class TrainMLAtractividadRequest(BaseModel):
    ultimos_anios: int = 3
    epochs: int = 800
    lr: float = 0.02


class AtractividadMLV2Request(BaseModel):
    proyecto_id: int
    monto_oferta: float = 0.0
    margen_pct: float = 0.0
    rubro: str = ""
    cliente: str = ""


class RequisitoDocumentalRequest(BaseModel):
    proyecto_id: int
    nombre: str
    categoria: str = "administrativo"
    estado: str = "pendiente"
    observaciones: str = ""


class RequisitoDocumentalUpdateRequest(BaseModel):
    estado: str = "pendiente"
    observaciones: str = ""


class LegalAceptacionRequest(BaseModel):
    accepted: bool = True
    terms_version: str = LEGAL_TERMS_CURRENT_VERSION
    accepted_at: Optional[str] = None
    accepted_source: str = "frontend-local"
    accepted_by: str = ""
    metadata: dict = Field(default_factory=dict)


class PlanCierreItemRequest(BaseModel):
    titulo: str
    prioridad: str = "media"
    owner: str = "equipo"
    estado: str = "pendiente"
    origen: str = "manual"
    metadata: dict = Field(default_factory=dict)


class PlanCierreItemUpdateRequest(BaseModel):
    prioridad: Optional[str] = None
    owner: Optional[str] = None
    estado: Optional[str] = None
    titulo: Optional[str] = None

# ─── Configuración IA en memoria (sesión) ─────────────────────────────────────
ia_config = {}
ML_DIR = EXPORTS_DIR / "ml"
ML_DIR.mkdir(exist_ok=True)
ML_ATRACTIVIDAD_MODEL_PATH = ML_DIR / "atractividad_ml_v2.json"


def _norm_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", (value or "").strip().lower()).strip("_")


def _to_float(v, default=0.0) -> float:
    if v is None:
        return default
    s = str(v).strip().replace(".", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        try:
            return float(v)
        except Exception:
            return default


def _to_bool(v) -> bool:
    s = str(v or "").strip().lower()
    return s in {"1", "true", "si", "sí", "yes", "y", "adjudicada", "ganada", "win"}


def _to_date(v):
    if v is None or str(v).strip() == "":
        return datetime.utcnow()
    if isinstance(v, datetime):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            pass
    try:
        return datetime.fromisoformat(s)
    except Exception:
        return datetime.utcnow()


def _qa_rules_for_tipo(tipo: str) -> list[dict]:
    t = (tipo or "").strip().lower()
    comunes = [
        {"key": "archivo_existe", "peso": 25, "label": "Archivo generado disponible"},
        {"key": "nombre_profesional", "peso": 10, "label": "Nombre de archivo profesional"},
        {"key": "tamano_archivo", "peso": 10, "label": "Tamaño mínimo de entregable"},
    ]
    if t in {"presupuesto_pdf", "presupuesto_excel"}:
        return comunes + [
            {"key": "estructura_financiera", "peso": 20, "label": "Estructura financiera mínima"},
            {"key": "metadatos_cliente", "peso": 10, "label": "Incluye cliente/proyecto"},
            {"key": "condiciones_comerciales", "peso": 15, "label": "Incluye validez/supuestos"},
            {"key": "contenido_extendido", "peso": 10, "label": "Contenido técnico suficiente"},
        ]
    if t in {"informe_pdf", "propuesta_pdf"}:
        return comunes + [
            {"key": "estructura_ejecutiva", "peso": 20, "label": "Estructura ejecutiva del contenido"},
            {"key": "riesgos_mitigaciones", "peso": 10, "label": "Incluye riesgos/mitigaciones"},
            {"key": "cierre_recomendacion", "peso": 15, "label": "Incluye cierre/recomendación"},
            {"key": "contenido_extendido", "peso": 10, "label": "Contenido técnico suficiente"},
        ]
    return comunes + [
        {"key": "contenido_extendido", "peso": 20, "label": "Contenido suficiente"},
        {"key": "estructura_ejecutiva", "peso": 15, "label": "Estructura ejecutiva"},
    ]


def _qa_check_document(doc: Documento) -> dict:
    path = Path(doc.archivo_generado or "")
    contenido = (doc.contenido or "").lower()
    nombre = (doc.nombre or "").lower()
    size = path.stat().st_size if path.exists() else 0
    tipo = (doc.tipo or "").lower()

    checks = {
        "archivo_existe": path.exists(),
        "nombre_profesional": any(x in nombre for x in ["presupuesto", "informe", "propuesta", "indice"]),
        "tamano_archivo": size >= 8_000,
        "estructura_financiera": any(x in contenido for x in ["subtotal", "iva", "total", "utilidades"]),
        "metadatos_cliente": any(x in contenido for x in ["cliente", "mandante", "proyecto"]),
        "condiciones_comerciales": any(x in contenido for x in ["validez", "supuesto", "plazo"]),
        "estructura_ejecutiva": any(x in contenido for x in ["resumen ejecutivo", "alcance", "metodología", "conclusiones"]),
        "riesgos_mitigaciones": any(x in contenido for x in ["riesgo", "mitigaci"]),
        "cierre_recomendacion": any(x in contenido for x in ["recomendación", "cierre", "conclusiones"]),
        "contenido_extendido": len(doc.contenido or "") >= 600,
    }

    rules = _qa_rules_for_tipo(tipo)
    score = 0
    hallazgos = []
    faltantes = []
    for r in rules:
        ok = bool(checks.get(r["key"], False))
        if ok:
            score += int(r["peso"])
            hallazgos.append(r["label"])
        else:
            faltantes.append(r["label"])

    recomendaciones_map = {
        "Archivo generado disponible": "Regenerar documento y verificar ruta de exportación.",
        "Nombre de archivo profesional": "Usar nomenclatura profesional (tipo + proyecto + fecha).",
        "Tamaño mínimo de entregable": "Agregar mayor profundidad técnica y anexos relevantes.",
        "Estructura financiera mínima": "Incluir subtotal, gastos generales, utilidades, neto, IVA y total.",
        "Incluye cliente/proyecto": "Agregar metadatos de cliente, proyecto y fecha en portada/encabezado.",
        "Incluye validez/supuestos": "Incorporar validez de oferta y supuestos comerciales explícitos.",
        "Estructura ejecutiva del contenido": "Agregar resumen ejecutivo, alcance, metodología y conclusiones.",
        "Incluye riesgos/mitigaciones": "Añadir matriz de riesgos y medidas de mitigación.",
        "Incluye cierre/recomendación": "Cerrar con recomendación ejecutiva y próximos pasos.",
        "Contenido técnico suficiente": "Aumentar detalle técnico y criterios de respaldo para cliente.",
    }
    recomendaciones = [recomendaciones_map.get(f, f"Mejorar: {f}") for f in faltantes]

    score = min(100, max(0, score))
    nivel = "pro" if score >= 85 else ("aceptable" if score >= 65 else "debil")
    return {
        "documento_id": doc.id,
        "tipo": doc.tipo,
        "nombre": doc.nombre,
        "score_qa": score,
        "nivel": nivel,
        "hallazgos": hallazgos,
        "faltantes": faltantes,
        "recomendaciones": recomendaciones,
        "archivo_existe": path.exists(),
        "archivo_size_bytes": size,
    }

# ═══════════════════════════════════════════════════════════════════════════════
# ENDPOINTS
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/")
async def root():
    return {"message": "DocPresupuestoAI API v1.0", "status": "online"}

@app.get("/health")
async def health():
    return {"status": "ok", "timestamp": datetime.now().isoformat()}

# ─── CONFIGURACIÓN IA ─────────────────────────────────────────────────────────
@app.post("/api/config-ia")
async def configurar_ia(config: ConfigIA):
    global ia_config
    ia_config = config.dict()
    return {"message": "Configuración de IA guardada", "provider": config.provider, "model": config.model}

@app.get("/api/config-ia")
async def obtener_config_ia():
    if ia_config:
        return {"configurado": True, "provider": ia_config.get("provider"), "model": ia_config.get("model")}
    return {"configurado": False}

# ─── PROYECTOS ────────────────────────────────────────────────────────────────
@app.get("/api/proyectos")
async def listar_proyectos(db: Session = Depends(get_db)):
    proyectos = db.query(Proyecto).order_by(Proyecto.fecha_creacion.desc()).all()
    return [{
        "id": p.id,
        "nombre": p.nombre,
        "codigo_licitacion": p.codigo_licitacion or "",
        "cliente": p.cliente,
        "estado": p.estado,
        "fecha_creacion": p.fecha_creacion.isoformat() if p.fecha_creacion else "",
        "tiene_datos": bool(p.datos_extraidos)
    } for p in proyectos]

@app.get("/api/proyectos/{proyecto_id}")
async def obtener_proyecto(proyecto_id: int, db: Session = Depends(get_db)):
    proyecto = db.query(Proyecto).filter(Proyecto.id == proyecto_id).first()
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    return {
        "id": proyecto.id,
        "nombre": proyecto.nombre,
        "codigo_licitacion": proyecto.codigo_licitacion or "",
        "cliente": proyecto.cliente,
        "descripcion": proyecto.descripcion,
        "estado": proyecto.estado,
        "datos_extraidos": proyecto.datos_extraidos,
        "fecha_creacion": proyecto.fecha_creacion.isoformat() if proyecto.fecha_creacion else ""
    }

@app.delete("/api/proyectos/{proyecto_id}")
async def eliminar_proyecto(proyecto_id: int, db: Session = Depends(get_db)):
    proyecto = db.query(Proyecto).filter(Proyecto.id == proyecto_id).first()
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    db.query(PlanCierreItem).filter(PlanCierreItem.proyecto_id == proyecto_id).delete()
    db.query(LegalAceptacionProyecto).filter(LegalAceptacionProyecto.proyecto_id == proyecto_id).delete()
    db.delete(proyecto)
    db.commit()
    return {"message": "Proyecto eliminado"}


@app.post("/api/demo/reset")
async def reset_demo_data(db: Session = Depends(get_db)):
    """
    Local-first helper: clears operational data for clean demos.
    """
    db.query(Documento).delete()
    db.query(ItemPresupuesto).delete()
    db.query(EvidenciaRequisito).delete()
    db.query(RequisitoDocumental).delete()
    db.query(PrediccionAdjudicacion).delete()
    db.query(OfertaLicitacion).delete()
    db.query(HistoricoLicitacion).delete()
    db.query(LegalAceptacionProyecto).delete()
    db.query(PlanCierreItem).delete()
    db.query(Proyecto).delete()
    db.commit()
    return {"message": "Datos demo reiniciados"}


@app.get("/api/proyectos/{proyecto_id}/legal-aceptacion")
async def obtener_aceptacion_legal_proyecto(proyecto_id: int, db: Session = Depends(get_db)):
    proyecto = db.query(Proyecto).filter(Proyecto.id == proyecto_id).first()
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    rec = (
        db.query(LegalAceptacionProyecto)
        .filter(LegalAceptacionProyecto.proyecto_id == proyecto_id)
        .order_by(LegalAceptacionProyecto.accepted_at.desc(), LegalAceptacionProyecto.id.desc())
        .first()
    )
    if not rec:
        return {
            "proyecto_id": proyecto_id,
            "aceptado": False,
            "accepted_at": None,
            "terms_version": LEGAL_TERMS_CURRENT_VERSION,
            "accepted_source": "",
            "accepted_by": "",
            "metadata": {},
        }
    return {
        "proyecto_id": proyecto_id,
        "aceptado": bool(rec.accepted),
        "accepted_at": rec.accepted_at.isoformat() if rec.accepted_at else None,
        "terms_version": rec.terms_version or LEGAL_TERMS_CURRENT_VERSION,
        "accepted_source": rec.accepted_source or "",
        "accepted_by": rec.accepted_by or "",
        "metadata": rec.metadata_json or {},
    }


@app.post("/api/proyectos/{proyecto_id}/legal-aceptacion")
async def registrar_aceptacion_legal_proyecto(
    proyecto_id: int,
    req: LegalAceptacionRequest,
    db: Session = Depends(get_db),
):
    proyecto = db.query(Proyecto).filter(Proyecto.id == proyecto_id).first()
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    accepted_at = datetime.utcnow()
    if req.accepted_at:
        try:
            parsed = datetime.fromisoformat(req.accepted_at.replace("Z", "+00:00"))
            accepted_at = parsed.replace(tzinfo=None) if parsed.tzinfo else parsed
        except Exception:
            accepted_at = datetime.utcnow()

    rec = LegalAceptacionProyecto(
        proyecto_id=proyecto_id,
        accepted=bool(req.accepted),
        accepted_at=accepted_at,
        terms_version=(req.terms_version or LEGAL_TERMS_CURRENT_VERSION).strip(),
        accepted_source=(req.accepted_source or "frontend-local").strip(),
        accepted_by=(req.accepted_by or "").strip(),
        metadata_json=req.metadata or {},
    )
    db.add(rec)
    db.commit()
    db.refresh(rec)
    return {
        "proyecto_id": proyecto_id,
        "aceptado": bool(rec.accepted),
        "accepted_at": rec.accepted_at.isoformat() if rec.accepted_at else None,
        "terms_version": rec.terms_version,
        "accepted_source": rec.accepted_source,
    }


@app.post("/api/demo/seed")
async def seed_demo_data(db: Session = Depends(get_db)):
    """
    Creates a representative local demo project with extracted-like data.
    """
    demo_codigo = "OT_26050101_DEMO_PULSO"
    existente = db.query(Proyecto).filter(Proyecto.codigo_licitacion == demo_codigo).first()
    if existente:
        return {
            "message": "Proyecto demo ya existente",
            "proyecto_id": existente.id,
            "codigo_licitacion": existente.codigo_licitacion or "",
        }

    datos_demo = {
        "proyecto": {
            "nombre": "Mantenimiento Integral Planta Norte",
            "mandante": "Industrial Demo S.A.",
            "moneda": "CLP",
            "modalidad": "Suma alzada",
            "plazo_ejecucion": "90 días",
            "ubicacion": "Santiago",
            "fecha_inicio": "2026-06-01",
            "fecha_termino": "2026-08-30",
            "descripcion": "Servicio integral de mantenimiento electromecánico y apoyo HSE.",
        },
        "requisitos_tecnicos": [
            "Plan de trabajo semanal",
            "Experiencia comprobable en mantenimiento industrial",
            "Cuadrilla certificada en trabajos críticos",
        ],
        "documentos_requeridos": [
            "Boleta de garantía de seriedad",
            "Certificado de cumplimiento previsional",
            "Declaración de seguridad y salud ocupacional",
        ],
        "partidas_presupuesto": [
            {"numero": "1", "descripcion": "Movilización y habilitación", "unidad": "gl", "cantidad": 1, "precio_unitario": 1850000},
            {"numero": "2", "descripcion": "Mano de obra especializada", "unidad": "hh", "cantidad": 540, "precio_unitario": 27500},
            {"numero": "3", "descripcion": "Materiales y consumibles", "unidad": "gl", "cantidad": 1, "precio_unitario": 4200000},
            {"numero": "4", "descripcion": "Supervisión y control de calidad", "unidad": "mes", "cantidad": 3, "precio_unitario": 1350000},
        ],
        "fechas_clave": {
            "Visita a terreno": "2026-05-10",
            "Consultas": "2026-05-12",
            "Cierre ofertas": "2026-05-20",
        },
    }

    proyecto_demo = Proyecto(
        nombre="Proyecto Demo - Mantenimiento Planta",
        codigo_licitacion=demo_codigo,
        cliente="Industrial Demo S.A.",
        descripcion=datos_demo["proyecto"]["descripcion"],
        archivo_base="DEMO_LOCAL",
        datos_extraidos=datos_demo,
        estado="activo",
    )
    db.add(proyecto_demo)
    db.commit()
    db.refresh(proyecto_demo)

    return {
        "message": "Proyecto demo creado",
        "proyecto_id": proyecto_demo.id,
        "codigo_licitacion": proyecto_demo.codigo_licitacion or "",
    }

# ─── SUBIR DOCUMENTO BASE ─────────────────────────────────────────────────────
@app.post("/api/subir-documento")
async def subir_documento(
    file: UploadFile = File(...),
    nombre_proyecto: str = Form(...),
    codigo_licitacion: str = Form(...),
    cliente: str = Form(default=""),
    provider: str = Form(default="openai"),
    api_key: str = Form(...),
    model: str = Form(default=""),
    db: Session = Depends(get_db)
):
    codigo_normalizado = _normalizar_codigo_licitacion(codigo_licitacion)
    if not CODIGO_LICITACION_REGEX.match(codigo_normalizado):
        raise HTTPException(
            status_code=400,
            detail="Codigo OT/Licitacion invalido. Usa formato OT_XXXXXXXX_CLIENTE (ej: OT_25086007_POLPAICO)",
        )

    # Guardar archivo
    ext = Path(file.filename).suffix
    filename = f"{uuid.uuid4()}{ext}"
    file_path = UPLOADS_DIR / filename
    
    with open(file_path, "wb") as f:
        shutil.copyfileobj(file.file, f)
    
    # Extraer texto
    try:
        texto = extract_text(str(file_path))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error al procesar archivo: {str(e)}")
    
    if not texto.strip():
        raise HTTPException(status_code=400, detail="No se pudo extraer texto del documento")
    
    # Analizar con IA
    try:
        engine = AIEngine(provider=provider, api_key=api_key, model=model)
        datos_extraidos = engine.analizar_bases(texto)
    except Exception as e:
        datos_extraidos = {"error": str(e), "texto_raw": texto[:2000]}
    
    # Guardar en BD
    proyecto = Proyecto(
        nombre=nombre_proyecto,
        codigo_licitacion=codigo_normalizado,
        cliente=cliente,
        descripcion=datos_extraidos.get("proyecto", {}).get("descripcion", ""),
        archivo_base=str(file_path),
        datos_extraidos=datos_extraidos
    )
    db.add(proyecto)
    db.commit()
    db.refresh(proyecto)
    
    return {
        "proyecto_id": proyecto.id,
        "nombre": proyecto.nombre,
        "datos_extraidos": datos_extraidos,
        "texto_extraido_chars": len(texto),
        "message": "Documento analizado exitosamente"
    }

# ─── GENERAR DOCUMENTOS ───────────────────────────────────────────────────────
@app.post("/api/generar-documento")
async def generar_documento(req: GenerarDocumentoRequest, db: Session = Depends(get_db)):
    proyecto = db.query(Proyecto).filter(Proyecto.id == req.proyecto_id).first()
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    datos_extraidos = proyecto.datos_extraidos or {}
    
    # Leer texto original
    texto_bases = ""
    if proyecto.archivo_base and Path(proyecto.archivo_base).exists():
        try:
            texto_bases = extract_text(proyecto.archivo_base)
        except:
            texto_bases = ""
    
    quality_mode = (req.quality_mode or "standard").strip().lower()
    engine = AIEngine(
        provider=req.config_ia.provider,
        api_key=req.config_ia.api_key,
        model=req.config_ia.model
    )
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre_safe = proyecto.nombre.replace(" ", "_")[:30]
    
    output_path = ""
    contenido = ""
    
    try:
        if req.tipo == "presupuesto_pdf":
            datos_ppto = engine.generar_presupuesto(datos_extraidos, texto_bases, quality_mode=quality_mode)
            filename = f"Presupuesto_{nombre_safe}_{timestamp}.pdf"
            output_path = str(EXPORTS_DIR / filename)
            generar_presupuesto_pdf(datos_ppto, output_path)
            contenido = json.dumps(datos_ppto)
            
        elif req.tipo == "presupuesto_excel":
            datos_ppto = engine.generar_presupuesto(datos_extraidos, texto_bases, quality_mode=quality_mode)
            filename = f"Presupuesto_{nombre_safe}_{timestamp}.xlsx"
            output_path = str(EXPORTS_DIR / filename)
            generar_presupuesto_excel(datos_ppto, output_path)
            contenido = json.dumps(datos_ppto)
            
        elif req.tipo == "informe_pdf":
            informe_md = engine.generar_informe_tecnico(datos_extraidos, texto_bases, quality_mode=quality_mode)
            filename = f"Informe_Tecnico_{nombre_safe}_{timestamp}.pdf"
            output_path = str(EXPORTS_DIR / filename)
            generar_informe_pdf(informe_md, datos_extraidos, output_path)
            contenido = informe_md
            
        elif req.tipo == "propuesta_pdf":
            propuesta_md = engine.generar_propuesta_tecnica(datos_extraidos, texto_bases, quality_mode=quality_mode)
            filename = f"Propuesta_Tecnica_{nombre_safe}_{timestamp}.pdf"
            output_path = str(EXPORTS_DIR / filename)
            generar_informe_pdf(propuesta_md, datos_extraidos, output_path)
            contenido = propuesta_md
        
        else:
            raise HTTPException(status_code=400, detail=f"Tipo de documento no válido: {req.tipo}")
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generando documento: {str(e)}")
    
    # Guardar en BD
    doc = Documento(
        proyecto_id=proyecto.id,
        tipo=req.tipo,
        nombre=filename,
        contenido=contenido[:5000],
        archivo_generado=output_path
    )
    db.add(doc)
    db.commit()
    db.refresh(doc)
    
    return {
        "documento_id": doc.id,
        "filename": filename,
        "tipo": req.tipo,
        "download_url": f"/api/descargar/{doc.id}",
        "message": "Documento generado exitosamente"
    }

# ─── DESCARGAR DOCUMENTO ──────────────────────────────────────────────────────
@app.get("/api/descargar/{documento_id}")
async def descargar_documento(documento_id: int, db: Session = Depends(get_db)):
    doc = db.query(Documento).filter(Documento.id == documento_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    if not Path(doc.archivo_generado).exists():
        raise HTTPException(status_code=404, detail="Archivo no encontrado en disco")
    if doc.tipo in PREFLIGHT_BLOCKED_DOWNLOAD_TYPES:
        preflight = await ejecutar_preflight_entrega(doc.proyecto_id, db)
        if preflight.get("estado_final") == "no_apto":
            return JSONResponse(
                status_code=409,
                content={
                    "message": "Descarga bloqueada por preflight NO APTO. Ejecuta plan de cierre.",
                    "estado_final": preflight.get("estado_final"),
                    "score_global": preflight.get("score_global"),
                    "plan_cierre": preflight.get("plan_cierre", []),
                    "brechas_criticas": preflight.get("brechas_criticas", []),
                },
            )
    
    return FileResponse(
        path=doc.archivo_generado,
        filename=doc.nombre,
        media_type="application/octet-stream"
    )

# ─── LISTAR DOCUMENTOS DE PROYECTO ───────────────────────────────────────────
@app.get("/api/proyectos/{proyecto_id}/documentos")
async def listar_documentos(proyecto_id: int, db: Session = Depends(get_db)):
    docs = db.query(Documento).filter(Documento.proyecto_id == proyecto_id).all()
    return [{
        "id": d.id,
        "proyecto_id": d.proyecto_id,
        "tipo": d.tipo,
        "nombre": d.nombre,
        "fecha_creacion": d.fecha_creacion.isoformat() if d.fecha_creacion else "",
        "download_url": f"/api/descargar/{d.id}"
    } for d in docs]


@app.get("/api/documentos/{documento_id}/qa")
async def evaluar_qa_documento(documento_id: int, db: Session = Depends(get_db)):
    doc = db.query(Documento).filter(Documento.id == documento_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    result = _qa_check_document(doc)
    result["mensaje"] = (
        "Documento en estándar pro de entrega."
        if result["score_qa"] >= 85
        else "Documento con brechas; revisar faltantes antes de envío a cliente."
    )
    return result


@app.get("/api/proyectos/{proyecto_id}/documentos/qa-resumen")
async def resumen_qa_documentos_proyecto(proyecto_id: int, db: Session = Depends(get_db)):
    docs = db.query(Documento).filter(Documento.proyecto_id == proyecto_id).all()
    if not docs:
        return {
            "proyecto_id": proyecto_id,
            "documentos": [],
            "score_promedio": 0,
            "nivel": "sin_documentos",
            "mensaje": "No hay documentos para evaluar.",
        }
    evaluaciones = [_qa_check_document(d) for d in docs]
    promedio = round(sum(e["score_qa"] for e in evaluaciones) / len(evaluaciones), 2)
    nivel = "pro" if promedio >= 85 else ("aceptable" if promedio >= 65 else "debil")
    return {
        "proyecto_id": proyecto_id,
        "documentos": evaluaciones,
        "score_promedio": promedio,
        "nivel": nivel,
        "mensaje": (
            "Portafolio documental en estándar pro."
            if promedio >= 85
            else "Portafolio documental con brechas de calidad."
        ),
    }


@app.get("/api/proyectos/{proyecto_id}/entrega-readiness")
async def evaluar_entrega_readiness(proyecto_id: int, db: Session = Depends(get_db)):
    """
    Evalúa si el proyecto está listo para entrega profesional a cliente.
    Devuelve score, nivel y brechas accionables.
    """
    proyecto = db.query(Proyecto).filter(Proyecto.id == proyecto_id).first()
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")

    docs = db.query(Documento).filter(Documento.proyecto_id == proyecto_id).all()
    requisitos = db.query(RequisitoDocumental).filter(RequisitoDocumental.proyecto_id == proyecto_id).all()
    evidencias = db.query(EvidenciaRequisito).filter(EvidenciaRequisito.proyecto_id == proyecto_id).all()
    predicciones = db.query(PrediccionAdjudicacion).filter(PrediccionAdjudicacion.proyecto_id == proyecto_id).all()
    ofertas = db.query(OfertaLicitacion).filter(OfertaLicitacion.proyecto_id == proyecto_id).all()

    tipos_docs = {d.tipo for d in docs}
    tiene_entregables_base = any(t in tipos_docs for t in {"presupuesto_pdf", "presupuesto_excel", "informe_pdf", "propuesta_pdf"})
    tiene_indice_documental = any(t in tipos_docs for t in {"indice_documental_pdf", "indice_documental_excel"})
    total_requisitos = len(requisitos)
    requisitos_cumplidos = sum(1 for r in requisitos if (r.estado or "").strip().lower() == "cumplido")
    cumplimiento_requisitos_pct = round((requisitos_cumplidos / total_requisitos) * 100, 2) if total_requisitos else 0.0
    tiene_evidencias = len(evidencias) > 0
    tiene_analitica_comercial = len(predicciones) > 0 and len(ofertas) > 0
    tiene_datos_ia = bool(proyecto.datos_extraidos)

    score = 0
    brechas = []

    if tiene_datos_ia:
        score += 20
    else:
        brechas.append("Falta análisis base de IA del proyecto.")

    if tiene_entregables_base:
        score += 25
    else:
        brechas.append("No hay entregables base (presupuesto/informe/propuesta) generados.")

    if tiene_indice_documental:
        score += 15
    else:
        brechas.append("Falta índice documental exportado (PDF o Excel).")

    if cumplimiento_requisitos_pct >= 80:
        score += 20
    elif cumplimiento_requisitos_pct >= 50:
        score += 10
        brechas.append("Cumplimiento documental parcial; elevar requisitos cumplidos sobre 80%.")
    else:
        brechas.append("Cumplimiento documental bajo; completar requisitos críticos.")

    if tiene_evidencias:
        score += 10
    else:
        brechas.append("No hay evidencias documentales cargadas.")

    if tiene_analitica_comercial:
        score += 10
    else:
        brechas.append("Falta analítica comercial (ofertas y predicciones) para respaldo ejecutivo.")

    nivel = "pro" if score >= 85 else ("avanzado" if score >= 65 else "basico")
    recomendado_para_cliente = score >= 75

    return {
        "proyecto_id": proyecto_id,
        "score_readiness": score,
        "nivel": nivel,
        "recomendado_para_cliente": recomendado_para_cliente,
        "metricas": {
            "tiene_datos_ia": tiene_datos_ia,
            "entregables_generados": len(docs),
            "tiene_entregables_base": tiene_entregables_base,
            "tiene_indice_documental": tiene_indice_documental,
            "total_requisitos": total_requisitos,
            "requisitos_cumplidos": requisitos_cumplidos,
            "cumplimiento_requisitos_pct": cumplimiento_requisitos_pct,
            "evidencias": len(evidencias),
            "ofertas": len(ofertas),
            "predicciones": len(predicciones),
        },
        "brechas": brechas,
        "mensaje": (
            "Proyecto listo para entrega cliente."
            if recomendado_para_cliente
            else "Proyecto aún no cumple estándar de entrega cliente; revisar brechas."
        ),
    }


@app.get("/api/proyectos/{proyecto_id}/preflight-entrega")
async def ejecutar_preflight_entrega(proyecto_id: int, db: Session = Depends(get_db)):
    """
    Veredicto final previo a envío a cliente:
    combina readiness general + QA documental.
    """
    proyecto = db.query(Proyecto).filter(Proyecto.id == proyecto_id).first()
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")

    # Reutilizar lógica de readiness.
    docs = db.query(Documento).filter(Documento.proyecto_id == proyecto_id).all()
    requisitos = db.query(RequisitoDocumental).filter(RequisitoDocumental.proyecto_id == proyecto_id).all()
    evidencias = db.query(EvidenciaRequisito).filter(EvidenciaRequisito.proyecto_id == proyecto_id).all()
    predicciones = db.query(PrediccionAdjudicacion).filter(PrediccionAdjudicacion.proyecto_id == proyecto_id).all()
    ofertas = db.query(OfertaLicitacion).filter(OfertaLicitacion.proyecto_id == proyecto_id).all()

    tipos_docs = {d.tipo for d in docs}
    tiene_entregables_base = any(t in tipos_docs for t in {"presupuesto_pdf", "presupuesto_excel", "informe_pdf", "propuesta_pdf"})
    tiene_indice_documental = any(t in tipos_docs for t in {"indice_documental_pdf", "indice_documental_excel"})
    total_requisitos = len(requisitos)
    requisitos_cumplidos = sum(1 for r in requisitos if (r.estado or "").strip().lower() == "cumplido")
    cumplimiento_requisitos_pct = round((requisitos_cumplidos / total_requisitos) * 100, 2) if total_requisitos else 0.0
    tiene_evidencias = len(evidencias) > 0
    tiene_analitica_comercial = len(predicciones) > 0 and len(ofertas) > 0
    tiene_datos_ia = bool(proyecto.datos_extraidos)

    score_readiness = 0
    brechas_readiness = []
    if tiene_datos_ia:
        score_readiness += 20
    else:
        brechas_readiness.append("Falta análisis base de IA del proyecto.")
    if tiene_entregables_base:
        score_readiness += 25
    else:
        brechas_readiness.append("No hay entregables base (presupuesto/informe/propuesta) generados.")
    if tiene_indice_documental:
        score_readiness += 15
    else:
        brechas_readiness.append("Falta índice documental exportado (PDF o Excel).")
    if cumplimiento_requisitos_pct >= 80:
        score_readiness += 20
    elif cumplimiento_requisitos_pct >= 50:
        score_readiness += 10
        brechas_readiness.append("Cumplimiento documental parcial; elevar requisitos cumplidos sobre 80%.")
    else:
        brechas_readiness.append("Cumplimiento documental bajo; completar requisitos críticos.")
    if tiene_evidencias:
        score_readiness += 10
    else:
        brechas_readiness.append("No hay evidencias documentales cargadas.")
    if tiene_analitica_comercial:
        score_readiness += 10
    else:
        brechas_readiness.append("Falta analítica comercial (ofertas y predicciones) para respaldo ejecutivo.")

    qa_docs = [_qa_check_document(d) for d in docs]
    score_qa = round(sum(x["score_qa"] for x in qa_docs) / len(qa_docs), 2) if qa_docs else 0.0
    legal_rec = (
        db.query(LegalAceptacionProyecto)
        .filter(LegalAceptacionProyecto.proyecto_id == proyecto_id)
        .order_by(LegalAceptacionProyecto.accepted_at.desc(), LegalAceptacionProyecto.id.desc())
        .first()
    )
    legal_aceptado = bool(legal_rec and legal_rec.accepted)
    legal_terms_version = legal_rec.terms_version if legal_rec else LEGAL_TERMS_CURRENT_VERSION
    legal_al_dia = legal_aceptado and legal_terms_version == LEGAL_TERMS_CURRENT_VERSION

    brechas_qa = []
    plan_cierre = []
    for ev in qa_docs:
        if ev["score_qa"] < PRELIGHT_QA_MIN_SCORE:
            brechas_qa.append(f"{ev['nombre']} ({ev['score_qa']}/100)")
            for rec in ev.get("recomendaciones", [])[:2]:
                if rec not in plan_cierre:
                    plan_cierre.append(rec)

    docs_por_tipo = {}
    for ev in qa_docs:
        tipo_doc = ev.get("tipo", "")
        docs_por_tipo.setdefault(tipo_doc, []).append(ev)
    mejor_score_por_tipo = {
        t: max((x.get("score_qa", 0) for x in evals), default=0)
        for t, evals in docs_por_tipo.items()
    }

    cobertura_items = []
    brechas_cobertura = []
    faltantes_criticos = []
    for grupo, tipos_grupo in PREFLIGHT_REQUIRED_DOC_GROUPS.items():
        presentes = [t for t in tipos_grupo if t in tipos_docs]
        best_score = max((mejor_score_por_tipo.get(t, 0) for t in presentes), default=0)
        qa_ok = bool(presentes) and best_score >= PRELIGHT_QA_MIN_SCORE
        item = {
            "grupo": grupo,
            "tipos_validos": tipos_grupo,
            "tipos_presentes": presentes,
            "presente": bool(presentes),
            "score_qa_max": best_score,
            "qa_ok": qa_ok,
        }
        cobertura_items.append(item)
        if not presentes:
            faltantes_criticos.append(grupo)
            brechas_cobertura.append(f"Falta entregable obligatorio del grupo '{grupo}'.")
            plan_cierre.append(f"Generar documento obligatorio del grupo '{grupo}'.")
        elif not qa_ok:
            brechas_cobertura.append(
                f"Grupo '{grupo}' bajo estándar QA ({best_score}/100, mínimo {PRELIGHT_QA_MIN_SCORE})."
            )
            plan_cierre.append(
                f"Regenerar o mejorar documento del grupo '{grupo}' hasta QA >= {PRELIGHT_QA_MIN_SCORE}."
            )

    grupos_ok = sum(1 for x in cobertura_items if x["qa_ok"])
    cobertura_pct = round((grupos_ok / len(cobertura_items)) * 100, 2) if cobertura_items else 0.0

    score_global = round((score_readiness * 0.55) + (score_qa * 0.45), 2)
    brechas_criticas = brechas_readiness + brechas_qa + brechas_cobertura
    if not legal_al_dia:
        brechas_criticas.append("Falta aceptación legal vigente del proyecto.")
        plan_cierre.insert(0, "Registrar aceptación legal del proyecto en versión vigente de términos.")

    if score_global >= 85 and len(brechas_criticas) <= 2:
        estado_final = "apto"
    elif score_global >= 70:
        estado_final = "condicional"
    else:
        estado_final = "no_apto"

    if estado_final == "apto" and not legal_al_dia:
        estado_final = "condicional"
    if faltantes_criticos:
        estado_final = "no_apto"
    elif estado_final == "apto" and cobertura_pct < 100:
        estado_final = "condicional"

    apto_para_cliente = estado_final == "apto"
    if not plan_cierre and not apto_para_cliente:
        plan_cierre = [
            "Regenerar documentos críticos en modo PRO.",
            "Completar evidencias y elevar cumplimiento documental sobre 80%.",
            "Recalcular predicción y bid leveling para respaldo ejecutivo.",
        ]
    plan_cierre_dedup = []
    for accion in plan_cierre:
        if accion not in plan_cierre_dedup:
            plan_cierre_dedup.append(accion)

    return {
        "proyecto_id": proyecto_id,
        "estado_final": estado_final,
        "apto_para_cliente": apto_para_cliente,
        "score_global": score_global,
        "score_readiness": score_readiness,
        "score_qa": score_qa,
        "cobertura_documental": {
            "grupos_requeridos": len(cobertura_items),
            "grupos_ok": grupos_ok,
            "cumplimiento_pct": cobertura_pct,
            "faltantes_criticos": faltantes_criticos,
            "detalle": cobertura_items,
            "qa_min_score": PRELIGHT_QA_MIN_SCORE,
        },
        "legal": {
            "aceptado": legal_aceptado,
            "al_dia": legal_al_dia,
            "terms_version_actual": LEGAL_TERMS_CURRENT_VERSION,
            "terms_version_aceptada": legal_terms_version,
            "accepted_at": legal_rec.accepted_at.isoformat() if legal_rec and legal_rec.accepted_at else None,
        },
        "brechas_criticas": brechas_criticas,
        "plan_cierre": plan_cierre_dedup[:8],
        "mensaje": (
            "Proyecto apto para entrega cliente."
            if apto_para_cliente
            else "Proyecto con brechas para entrega; ejecutar plan de cierre."
        ),
    }


def _prioridad_desde_accion(texto: str) -> str:
    t = (texto or "").lower()
    if any(x in t for x in ["crític", "bloque", "obligatorio", "legal", "falt"]):
        return "alta"
    if any(x in t for x in ["qa", "evidencia", "requisito"]):
        return "media"
    return "baja"


@app.get("/api/proyectos/{proyecto_id}/plan-cierre")
async def listar_plan_cierre(proyecto_id: int, db: Session = Depends(get_db)):
    proyecto = db.query(Proyecto).filter(Proyecto.id == proyecto_id).first()
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    items = (
        db.query(PlanCierreItem)
        .filter(PlanCierreItem.proyecto_id == proyecto_id)
        .order_by(PlanCierreItem.fecha_creacion.desc())
        .all()
    )
    return {
        "proyecto_id": proyecto_id,
        "items": [{
            "id": i.id,
            "titulo": i.titulo,
            "prioridad": i.prioridad,
            "owner": i.owner,
            "estado": i.estado,
            "origen": i.origen,
            "metadata": i.metadata_json or {},
            "fecha_creacion": i.fecha_creacion.isoformat() if i.fecha_creacion else "",
            "fecha_actualizacion": i.fecha_actualizacion.isoformat() if i.fecha_actualizacion else "",
        } for i in items]
    }


@app.post("/api/proyectos/{proyecto_id}/plan-cierre")
async def crear_item_plan_cierre(proyecto_id: int, req: PlanCierreItemRequest, db: Session = Depends(get_db)):
    proyecto = db.query(Proyecto).filter(Proyecto.id == proyecto_id).first()
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    item = PlanCierreItem(
        proyecto_id=proyecto_id,
        titulo=(req.titulo or "").strip(),
        prioridad=(req.prioridad or "media").strip().lower(),
        owner=(req.owner or "equipo").strip(),
        estado=(req.estado or "pendiente").strip().lower(),
        origen=(req.origen or "manual").strip().lower(),
        metadata_json=req.metadata or {},
        fecha_actualizacion=datetime.utcnow(),
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    return {"item_id": item.id, "message": "Item de cierre creado"}


@app.post("/api/proyectos/{proyecto_id}/plan-cierre/sincronizar-preflight")
async def sincronizar_plan_cierre_desde_preflight(proyecto_id: int, db: Session = Depends(get_db)):
    preflight = await ejecutar_preflight_entrega(proyecto_id, db)
    candidatos = []
    for x in (preflight.get("plan_cierre") or []):
        if x and isinstance(x, str):
            candidatos.append(x.strip())
    for x in (preflight.get("brechas_criticas") or []):
        if x and isinstance(x, str):
            candidatos.append(f"Resolver brecha: {x.strip()}")

    existentes = db.query(PlanCierreItem).filter(PlanCierreItem.proyecto_id == proyecto_id).all()
    firmas_existentes = {(e.titulo or "").strip().lower() for e in existentes}
    creados = 0
    for titulo in candidatos:
        firma = titulo.lower()
        if firma in firmas_existentes:
            continue
        item = PlanCierreItem(
            proyecto_id=proyecto_id,
            titulo=titulo,
            prioridad=_prioridad_desde_accion(titulo),
            owner="equipo",
            estado="pendiente",
            origen="preflight",
            metadata_json={"estado_preflight": preflight.get("estado_final", "")},
            fecha_actualizacion=datetime.utcnow(),
        )
        db.add(item)
        firmas_existentes.add(firma)
        creados += 1
    db.commit()
    return {
        "proyecto_id": proyecto_id,
        "creados": creados,
        "estado_preflight": preflight.get("estado_final"),
        "message": "Plan de cierre sincronizado desde preflight",
    }


@app.put("/api/plan-cierre/{item_id}")
async def actualizar_item_plan_cierre(item_id: int, req: PlanCierreItemUpdateRequest, db: Session = Depends(get_db)):
    item = db.query(PlanCierreItem).filter(PlanCierreItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Item de cierre no encontrado")
    if req.titulo is not None:
        item.titulo = req.titulo.strip()
    if req.owner is not None:
        item.owner = req.owner.strip()
    if req.prioridad is not None:
        item.prioridad = req.prioridad.strip().lower()
    if req.estado is not None:
        item.estado = req.estado.strip().lower()
    item.fecha_actualizacion = datetime.utcnow()
    db.commit()
    return {"message": "Item de cierre actualizado"}

# ─── CONSULTA LIBRE ───────────────────────────────────────────────────────────
@app.post("/api/consulta")
async def consulta_libre(req: ConsultaRequest, db: Session = Depends(get_db)):
    proyecto = db.query(Proyecto).filter(Proyecto.id == req.proyecto_id).first()
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    contexto = json.dumps(proyecto.datos_extraidos or {}, ensure_ascii=False)
    
    try:
        engine = AIEngine(
            provider=req.config_ia.provider,
            api_key=req.config_ia.api_key,
            model=req.config_ia.model
        )
        respuesta = engine.consulta_libre(req.pregunta, contexto)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    
    return {"respuesta": respuesta, "pregunta": req.pregunta}

# ─── DATOS EXTRAÍDOS ─────────────────────────────────────────────────────────
@app.get("/api/proyectos/{proyecto_id}/datos")
async def obtener_datos_extraidos(proyecto_id: int, db: Session = Depends(get_db)):
    proyecto = db.query(Proyecto).filter(Proyecto.id == proyecto_id).first()
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    return proyecto.datos_extraidos or {}

@app.put("/api/proyectos/{proyecto_id}/datos")
async def actualizar_datos(proyecto_id: int, datos: dict, db: Session = Depends(get_db)):
    proyecto = db.query(Proyecto).filter(Proyecto.id == proyecto_id).first()
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    proyecto.datos_extraidos = datos
    proyecto.fecha_actualizacion = datetime.utcnow()
    db.commit()
    return {"message": "Datos actualizados"}


# ─── DOCUMENTACION REQUERIDA ──────────────────────────────────────────────────
@app.post("/api/requisitos/sincronizar/{proyecto_id}")
async def sincronizar_requisitos_desde_ia(proyecto_id: int, db: Session = Depends(get_db)):
    proyecto = db.query(Proyecto).filter(Proyecto.id == proyecto_id).first()
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")

    datos = proyecto.datos_extraidos or {}
    candidatos = []
    candidatos += [("administrativo", x) for x in (datos.get("requisitos_administrativos") or [])]
    candidatos += [("tecnico", x) for x in (datos.get("requisitos_tecnicos") or [])]
    candidatos += [("administrativo", x) for x in (datos.get("documentos_requeridos") or [])]

    creados = 0
    for categoria, nombre in candidatos:
        nombre_limpio = str(nombre).strip()
        if not nombre_limpio:
            continue
        existente = (
            db.query(RequisitoDocumental)
            .filter(
                RequisitoDocumental.proyecto_id == proyecto_id,
                RequisitoDocumental.nombre == nombre_limpio,
            )
            .first()
        )
        if existente:
            continue
        nuevo = RequisitoDocumental(
            proyecto_id=proyecto_id,
            nombre=nombre_limpio,
            categoria=categoria,
            estado="pendiente",
            fuente="ia",
        )
        db.add(nuevo)
        creados += 1

    db.commit()
    return {"message": "Requisitos sincronizados", "creados": creados}


@app.get("/api/proyectos/{proyecto_id}/requisitos")
async def listar_requisitos_documentales(proyecto_id: int, db: Session = Depends(get_db)):
    requisitos = (
        db.query(RequisitoDocumental)
        .filter(RequisitoDocumental.proyecto_id == proyecto_id)
        .order_by(RequisitoDocumental.fecha_creacion.desc())
        .all()
    )
    evidencias = (
        db.query(EvidenciaRequisito)
        .filter(EvidenciaRequisito.proyecto_id == proyecto_id)
        .all()
    )
    conteo_evidencias = {}
    for e in evidencias:
        conteo_evidencias[e.requisito_id] = conteo_evidencias.get(e.requisito_id, 0) + 1

    return [
        {
            "id": r.id,
            "proyecto_id": r.proyecto_id,
            "nombre": r.nombre,
            "categoria": r.categoria,
            "seccion": _seccion_documental_por_categoria(r.categoria),
            "estado": r.estado,
            "observaciones": r.observaciones or "",
            "fuente": r.fuente or "manual",
            "carpeta_objetivo": str(_carpeta_requisito(r.proyecto_id, r.categoria, r.nombre).relative_to(BASE_DIR)),
            "evidencias_count": conteo_evidencias.get(r.id, 0),
            "fecha_creacion": r.fecha_creacion.isoformat() if r.fecha_creacion else "",
            "fecha_actualizacion": r.fecha_actualizacion.isoformat() if r.fecha_actualizacion else "",
        }
        for r in requisitos
    ]


@app.post("/api/requisitos")
async def crear_requisito_documental(req: RequisitoDocumentalRequest, db: Session = Depends(get_db)):
    proyecto = db.query(Proyecto).filter(Proyecto.id == req.proyecto_id).first()
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    if not req.nombre.strip():
        raise HTTPException(status_code=400, detail="Nombre de requisito es obligatorio")

    requisito = RequisitoDocumental(
        proyecto_id=req.proyecto_id,
        nombre=req.nombre.strip(),
        categoria=req.categoria or "administrativo",
        estado=req.estado or "pendiente",
        observaciones=req.observaciones or "",
        fuente="manual",
        fecha_actualizacion=datetime.utcnow(),
    )
    db.add(requisito)
    db.commit()
    db.refresh(requisito)
    return {"id": requisito.id, "message": "Requisito creado"}


@app.put("/api/requisitos/{requisito_id}")
async def actualizar_requisito_documental(
    requisito_id: int, req: RequisitoDocumentalUpdateRequest, db: Session = Depends(get_db)
):
    requisito = db.query(RequisitoDocumental).filter(RequisitoDocumental.id == requisito_id).first()
    if not requisito:
        raise HTTPException(status_code=404, detail="Requisito no encontrado")

    requisito.estado = req.estado or requisito.estado
    requisito.observaciones = req.observaciones or ""
    requisito.fecha_actualizacion = datetime.utcnow()
    db.commit()

    return {"message": "Requisito actualizado", "id": requisito.id}


@app.post("/api/requisitos/{requisito_id}/evidencia")
async def subir_evidencia_requisito(requisito_id: int, file: UploadFile = File(...), db: Session = Depends(get_db)):
    requisito = db.query(RequisitoDocumental).filter(RequisitoDocumental.id == requisito_id).first()
    if not requisito:
        raise HTTPException(status_code=404, detail="Requisito no encontrado")

    ext = Path(file.filename or "").suffix.lower()
    nombre_limpio = Path(file.filename or "").name or f"evidencia_{uuid.uuid4().hex[:8]}"
    carpeta = _carpeta_requisito(requisito.proyecto_id, requisito.categoria, requisito.nombre)
    carpeta.mkdir(parents=True, exist_ok=True)

    nombre_guardado = f"{uuid.uuid4().hex}_{nombre_limpio}"
    destino = carpeta / nombre_guardado
    with open(destino, "wb") as f:
        shutil.copyfileobj(file.file, f)

    ultimo_orden = (
        db.query(EvidenciaRequisito)
        .filter(EvidenciaRequisito.requisito_id == requisito_id)
        .order_by(EvidenciaRequisito.orden.desc())
        .first()
    )
    siguiente_orden = (ultimo_orden.orden + 1) if ultimo_orden else 1

    evidencia = EvidenciaRequisito(
        proyecto_id=requisito.proyecto_id,
        requisito_id=requisito.id,
        nombre_archivo=nombre_limpio,
        extension=ext,
        archivo_path=str(destino),
        carpeta_relativa=str(carpeta.relative_to(BASE_DIR)),
        orden=siguiente_orden,
    )
    db.add(evidencia)

    if requisito.estado == "pendiente":
        requisito.estado = "en_revision"
        requisito.fecha_actualizacion = datetime.utcnow()

    db.commit()
    db.refresh(evidencia)

    return {
        "id": evidencia.id,
        "requisito_id": evidencia.requisito_id,
        "nombre_archivo": evidencia.nombre_archivo,
        "carpeta_relativa": evidencia.carpeta_relativa,
        "orden": evidencia.orden,
        "message": "Evidencia subida",
    }


@app.get("/api/requisitos/{requisito_id}/evidencias")
async def listar_evidencias_requisito(requisito_id: int, db: Session = Depends(get_db)):
    evidencias = (
        db.query(EvidenciaRequisito)
        .filter(EvidenciaRequisito.requisito_id == requisito_id)
        .order_by(EvidenciaRequisito.orden.asc(), EvidenciaRequisito.fecha_creacion.asc())
        .all()
    )
    return [
        {
            "id": e.id,
            "requisito_id": e.requisito_id,
            "nombre_archivo": e.nombre_archivo,
            "extension": e.extension,
            "carpeta_relativa": e.carpeta_relativa,
            "orden": e.orden,
            "download_url": f"/api/evidencias/{e.id}/descargar",
            "fecha_creacion": e.fecha_creacion.isoformat() if e.fecha_creacion else "",
        }
        for e in evidencias
    ]


@app.get("/api/evidencias/{evidencia_id}/descargar")
async def descargar_evidencia(evidencia_id: int, db: Session = Depends(get_db)):
    evidencia = db.query(EvidenciaRequisito).filter(EvidenciaRequisito.id == evidencia_id).first()
    if not evidencia:
        raise HTTPException(status_code=404, detail="Evidencia no encontrada")
    if not Path(evidencia.archivo_path).exists():
        raise HTTPException(status_code=404, detail="Archivo de evidencia no encontrado")
    return FileResponse(
        path=evidencia.archivo_path,
        filename=evidencia.nombre_archivo,
        media_type="application/octet-stream",
    )


@app.get("/api/proyectos/{proyecto_id}/documentacion/arbol")
async def obtener_arbol_documentacion(proyecto_id: int, db: Session = Depends(get_db)):
    requisitos = (
        db.query(RequisitoDocumental)
        .filter(RequisitoDocumental.proyecto_id == proyecto_id)
        .order_by(RequisitoDocumental.categoria.asc(), RequisitoDocumental.nombre.asc())
        .all()
    )
    evidencias = (
        db.query(EvidenciaRequisito)
        .filter(EvidenciaRequisito.proyecto_id == proyecto_id)
        .order_by(EvidenciaRequisito.carpeta_relativa.asc(), EvidenciaRequisito.orden.asc())
        .all()
    )

    evidencias_por_req = {}
    for e in evidencias:
        evidencias_por_req.setdefault(e.requisito_id, []).append({
            "id": e.id,
            "nombre_archivo": e.nombre_archivo,
            "orden": e.orden,
            "carpeta_relativa": e.carpeta_relativa,
            "download_url": f"/api/evidencias/{e.id}/descargar",
        })

    secciones = {
        codigo: {"seccion": codigo, "categoria_base": categoria, "requisitos": []}
        for codigo, categoria in SECCIONES_DOCUMENTALES
    }

    for r in requisitos:
        seccion = _seccion_documental_por_categoria(r.categoria)
        secciones.setdefault(seccion, {"seccion": seccion, "categoria_base": "administrativo", "requisitos": []})
        secciones[seccion]["requisitos"].append({
            "requisito_id": r.id,
            "requisito": r.nombre,
            "categoria": r.categoria,
            "estado": r.estado,
            "carpeta_objetivo": str(_carpeta_requisito(r.proyecto_id, r.categoria, r.nombre).relative_to(BASE_DIR)),
            "evidencias": evidencias_por_req.get(r.id, []),
        })

    return {
        "proyecto_id": proyecto_id,
        "secciones": [secciones[codigo] for codigo, _ in SECCIONES_DOCUMENTALES],
    }


@app.post("/api/proyectos/{proyecto_id}/documentacion/exportar")
async def exportar_indice_documental(proyecto_id: int, formato: str = "excel", db: Session = Depends(get_db)):
    proyecto = db.query(Proyecto).filter(Proyecto.id == proyecto_id).first()
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    preflight = await ejecutar_preflight_entrega(proyecto_id, db)
    if preflight.get("estado_final") == "no_apto":
        return JSONResponse(
            status_code=409,
            content={
                "message": "Exportación bloqueada por preflight NO APTO. Ejecuta plan de cierre.",
                "estado_final": preflight.get("estado_final"),
                "score_global": preflight.get("score_global"),
                "plan_cierre": preflight.get("plan_cierre", []),
                "brechas_criticas": preflight.get("brechas_criticas", []),
            },
        )

    requisitos = (
        db.query(RequisitoDocumental)
        .filter(RequisitoDocumental.proyecto_id == proyecto_id)
        .order_by(RequisitoDocumental.categoria.asc(), RequisitoDocumental.nombre.asc())
        .all()
    )
    evidencias = (
        db.query(EvidenciaRequisito)
        .filter(EvidenciaRequisito.proyecto_id == proyecto_id)
        .order_by(EvidenciaRequisito.carpeta_relativa.asc(), EvidenciaRequisito.orden.asc())
        .all()
    )

    evidencias_por_req = {}
    for e in evidencias:
        evidencias_por_req.setdefault(e.requisito_id, []).append({
            "nombre_archivo": e.nombre_archivo,
            "orden": e.orden,
            "carpeta_relativa": e.carpeta_relativa,
            "download_url": f"/api/evidencias/{e.id}/descargar",
        })

    secciones = {
        codigo: {"seccion": codigo, "categoria_base": categoria, "requisitos": []}
        for codigo, categoria in SECCIONES_DOCUMENTALES
    }
    for r in requisitos:
        seccion = _seccion_documental_por_categoria(r.categoria)
        secciones[seccion]["requisitos"].append({
            "requisito_id": r.id,
            "requisito": r.nombre,
            "categoria": r.categoria,
            "estado": r.estado,
            "carpeta_objetivo": str(_carpeta_requisito(r.proyecto_id, r.categoria, r.nombre).relative_to(BASE_DIR)),
            "evidencias": evidencias_por_req.get(r.id, []),
        })
    secciones_ordenadas = [secciones[codigo] for codigo, _ in SECCIONES_DOCUMENTALES]

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre_safe = _slugify(proyecto.nombre)[:30]
    fecha_generacion = datetime.now().strftime("%d/%m/%Y %H:%M")
    codigo_licitacion = (proyecto.codigo_licitacion or "").strip()
    logo_candidates = [
        BASE_DIR / "templates" / "logo.png",
        BASE_DIR / "templates" / "logo.jpg",
        BASE_DIR / "templates" / "logo.jpeg",
    ]
    logo_path = ""
    for candidate in logo_candidates:
        if candidate.exists():
            logo_path = str(candidate)
            break
    fmt = (formato or "excel").strip().lower()
    if fmt not in {"excel", "pdf"}:
        raise HTTPException(status_code=400, detail="Formato no valido. Usa: excel o pdf")

    if fmt == "excel":
        filename = f"Indice_Documental_{nombre_safe}_{timestamp}.xlsx"
        output_path = str(EXPORTS_DIR / filename)
        generar_indice_documental_excel(
            proyecto_nombre=proyecto.nombre,
            secciones_ordenadas=secciones_ordenadas,
            output_path=output_path,
            cliente=proyecto.cliente or "",
            fecha_generacion=fecha_generacion,
            codigo_licitacion=codigo_licitacion,
        )
        tipo_doc = "indice_documental_excel"
    else:
        filename = f"Indice_Documental_{nombre_safe}_{timestamp}.pdf"
        output_path = str(EXPORTS_DIR / filename)
        generar_indice_documental_pdf(
            proyecto_nombre=proyecto.nombre,
            secciones_ordenadas=secciones_ordenadas,
            output_path=output_path,
            cliente=proyecto.cliente or "",
            fecha_generacion=fecha_generacion,
            logo_path=logo_path,
            codigo_licitacion=codigo_licitacion,
        )
        tipo_doc = "indice_documental_pdf"

    doc = Documento(
        proyecto_id=proyecto.id,
        tipo=tipo_doc,
        nombre=filename,
        contenido="Indice documental exportado",
        datos_json={"formato": fmt},
        archivo_generado=output_path,
    )
    db.add(doc)
    db.commit()
    db.refresh(doc)

    return {
        "documento_id": doc.id,
        "filename": filename,
        "tipo": tipo_doc,
        "download_url": f"/api/descargar/{doc.id}",
        "message": "Indice documental exportado exitosamente",
    }


# ─── PREDICCION DE ADJUDICACION ──────────────────────────────────────────────
@app.post("/api/prediccion/oferta")
async def crear_oferta_licitacion(req: OfertaLicitacionRequest, db: Session = Depends(get_db)):
    proyecto = db.query(Proyecto).filter(Proyecto.id == req.proyecto_id).first()
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")

    oferta = OfertaLicitacion(
        proyecto_id=req.proyecto_id,
        nombre=req.nombre,
        monto_oferta=req.monto_oferta,
        plazo_dias=req.plazo_dias,
        factores=req.factores or {},
        notas=req.notas,
    )
    db.add(oferta)
    db.commit()
    db.refresh(oferta)

    return {
        "oferta_id": oferta.id,
        "proyecto_id": oferta.proyecto_id,
        "nombre": oferta.nombre,
        "message": "Oferta registrada",
    }


@app.get("/api/proyectos/{proyecto_id}/ofertas")
async def listar_ofertas_proyecto(proyecto_id: int, db: Session = Depends(get_db)):
    ofertas = (
        db.query(OfertaLicitacion)
        .filter(OfertaLicitacion.proyecto_id == proyecto_id)
        .order_by(OfertaLicitacion.fecha_creacion.desc())
        .all()
    )
    return [
        {
            "id": o.id,
            "nombre": o.nombre,
            "monto_oferta": o.monto_oferta,
            "plazo_dias": o.plazo_dias,
            "factores": o.factores or {},
            "fecha_creacion": o.fecha_creacion.isoformat() if o.fecha_creacion else "",
        }
        for o in ofertas
    ]


@app.post("/api/prediccion/calcular/{oferta_id}")
async def calcular_prediccion_adjudicacion(oferta_id: int, db: Session = Depends(get_db)):
    oferta = db.query(OfertaLicitacion).filter(OfertaLicitacion.id == oferta_id).first()
    if not oferta:
        raise HTTPException(status_code=404, detail="Oferta no encontrada")

    proyecto = db.query(Proyecto).filter(Proyecto.id == oferta.proyecto_id).first()
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")

    resultado = calcular_prediccion(proyecto.datos_extraidos or {}, oferta.factores or {})

    pred = PrediccionAdjudicacion(
        proyecto_id=proyecto.id,
        oferta_id=oferta.id,
        score=resultado["score"],
        probabilidad=resultado["probabilidad_adjudicacion"],
        resultado_json=resultado,
        version_modelo=resultado.get("version_modelo", "rules-v1"),
    )
    db.add(pred)
    db.commit()
    db.refresh(pred)

    return {
        "prediccion_id": pred.id,
        "proyecto_id": pred.proyecto_id,
        "oferta_id": pred.oferta_id,
        "resultado": resultado,
    }


@app.get("/api/proyectos/{proyecto_id}/predicciones")
async def listar_predicciones_proyecto(proyecto_id: int, db: Session = Depends(get_db)):
    predicciones = (
        db.query(PrediccionAdjudicacion)
        .filter(PrediccionAdjudicacion.proyecto_id == proyecto_id)
        .order_by(PrediccionAdjudicacion.fecha_creacion.desc())
        .all()
    )
    ofertas = (
        db.query(OfertaLicitacion)
        .filter(OfertaLicitacion.proyecto_id == proyecto_id)
        .all()
    )
    oferta_nombre_por_id = {o.id: o.nombre for o in ofertas}

    return [
        {
            "id": p.id,
            "oferta_id": p.oferta_id,
            "oferta_nombre": oferta_nombre_por_id.get(p.oferta_id, f"Oferta #{p.oferta_id}"),
            "score": p.score,
            "probabilidad": p.probabilidad,
            "version_modelo": p.version_modelo,
            "fecha_creacion": p.fecha_creacion.isoformat() if p.fecha_creacion else "",
            "resultado": p.resultado_json or {},
        }
        for p in predicciones
    ]


@app.post("/api/prediccion/simular")
async def simular_prediccion(req: SimularPrediccionRequest, db: Session = Depends(get_db)):
    proyecto = db.query(Proyecto).filter(Proyecto.id == req.proyecto_id).first()
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    if not req.escenarios:
        raise HTTPException(status_code=400, detail="Debes enviar al menos un escenario")

    resultados = []
    for esc in req.escenarios[:3]:
        resultado = calcular_prediccion(proyecto.datos_extraidos or {}, esc.factores or {})
        resultados.append({
            "escenario": esc.nombre,
            "resultado": resultado,
        })

    resultados = sorted(resultados, key=lambda x: x["resultado"].get("score", 0), reverse=True)

    return {
        "proyecto_id": proyecto.id,
        "total_escenarios": len(resultados),
        "ranking": resultados,
    }


@app.post("/api/historico-licitaciones")
async def crear_historico_licitacion(req: HistoricoLicitacionRequest, db: Session = Depends(get_db)):
    if not (req.cliente or "").strip():
        raise HTTPException(status_code=400, detail="Cliente es obligatorio")
    fecha_cierre = datetime.utcnow()
    if (req.fecha_cierre or "").strip():
        try:
            fecha_cierre = datetime.fromisoformat(req.fecha_cierre)
        except Exception:
            raise HTTPException(status_code=400, detail="fecha_cierre invalida, usa formato ISO")

    h = HistoricoLicitacion(
        codigo_licitacion=(req.codigo_licitacion or "").strip().upper(),
        cliente=(req.cliente or "").strip(),
        rubro=(req.rubro or "").strip(),
        monto_ofertado=req.monto_ofertado or 0.0,
        margen_pct=req.margen_pct or 0.0,
        fue_adjudicada=bool(req.fue_adjudicada),
        fecha_cierre=fecha_cierre,
        observaciones=req.observaciones or "",
    )
    db.add(h)
    db.commit()
    db.refresh(h)
    return {"id": h.id, "message": "Historico registrado"}


@app.post("/api/historico-licitaciones/importar")
async def importar_historico_licitaciones(file: UploadFile = File(...), db: Session = Depends(get_db)):
    ext = Path(file.filename or "").suffix.lower()
    if ext not in {".csv", ".xlsx"}:
        raise HTTPException(status_code=400, detail="Formato no soportado. Usa CSV o XLSX")

    rows = []
    if ext == ".csv":
        content = await file.read()
        text = content.decode("utf-8", errors="ignore")
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            rows.append({(_norm_header(k)): v for k, v in (row or {}).items()})
    else:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        raw_headers = [cell.value for cell in ws[1]]
        headers = [_norm_header(str(h or "")) for h in raw_headers]
        for r in ws.iter_rows(min_row=2, values_only=True):
            rows.append({headers[i]: r[i] for i in range(len(headers))})

    alias = {
        "codigo": "codigo_licitacion",
        "ot": "codigo_licitacion",
        "codigo_ot": "codigo_licitacion",
        "codigo_licitacion": "codigo_licitacion",
        "cliente": "cliente",
        "mandante": "cliente",
        "rubro": "rubro",
        "categoria": "rubro",
        "monto": "monto_ofertado",
        "monto_ofertado": "monto_ofertado",
        "monto_oferta": "monto_ofertado",
        "margen": "margen_pct",
        "margen_pct": "margen_pct",
        "adjudicada": "fue_adjudicada",
        "fue_adjudicada": "fue_adjudicada",
        "resultado": "fue_adjudicada",
        "fecha": "fecha_cierre",
        "fecha_cierre": "fecha_cierre",
        "observaciones": "observaciones",
        "comentarios": "observaciones",
    }

    imported = 0
    skipped = 0
    errores = []
    for idx, row in enumerate(rows, start=2):
        canon = {}
        for k, v in row.items():
            mapped = alias.get(k, k)
            canon[mapped] = v

        cliente = str(canon.get("cliente") or "").strip()
        if not cliente:
            skipped += 1
            errores.append(f"Fila {idx}: cliente vacio")
            continue

        try:
            h = HistoricoLicitacion(
                codigo_licitacion=str(canon.get("codigo_licitacion") or "").strip().upper(),
                cliente=cliente,
                rubro=str(canon.get("rubro") or "").strip(),
                monto_ofertado=_to_float(canon.get("monto_ofertado"), 0.0),
                margen_pct=_to_float(canon.get("margen_pct"), 0.0),
                fue_adjudicada=_to_bool(canon.get("fue_adjudicada")),
                fecha_cierre=_to_date(canon.get("fecha_cierre")),
                observaciones=str(canon.get("observaciones") or "").strip(),
            )
            db.add(h)
            imported += 1
        except Exception as e:
            skipped += 1
            errores.append(f"Fila {idx}: {str(e)}")

    db.commit()
    return {
        "message": "Importacion completada",
        "archivo": file.filename,
        "importados": imported,
        "omitidos": skipped,
        "errores": errores[:30],
        "template_columnas": [
            "codigo_licitacion",
            "cliente",
            "rubro",
            "monto_ofertado",
            "margen_pct",
            "fue_adjudicada",
            "fecha_cierre",
            "observaciones",
        ],
    }


@app.get("/api/historico-licitaciones")
async def listar_historico_licitaciones(ultimos_anios: int = 3, db: Session = Depends(get_db)):
    limite_fecha = datetime.utcnow() - timedelta(days=max(1, ultimos_anios) * 365)
    historico = (
        db.query(HistoricoLicitacion)
        .filter(HistoricoLicitacion.fecha_cierre >= limite_fecha)
        .order_by(HistoricoLicitacion.fecha_cierre.desc())
        .all()
    )
    return [
        {
            "id": h.id,
            "codigo_licitacion": h.codigo_licitacion or "",
            "cliente": h.cliente or "",
            "rubro": h.rubro or "",
            "monto_ofertado": h.monto_ofertado or 0.0,
            "margen_pct": h.margen_pct or 0.0,
            "fue_adjudicada": bool(h.fue_adjudicada),
            "fecha_cierre": h.fecha_cierre.isoformat() if h.fecha_cierre else "",
            "observaciones": h.observaciones or "",
        }
        for h in historico
    ]


@app.post("/api/prediccion/atractividad")
async def evaluar_atractividad(req: AtractividadRequest, db: Session = Depends(get_db)):
    proyecto = db.query(Proyecto).filter(Proyecto.id == req.proyecto_id).first()
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")

    pred_actual = calcular_prediccion(proyecto.datos_extraidos or {}, req.factores or {})
    limite_fecha = datetime.utcnow() - timedelta(days=3 * 365)
    historico = (
        db.query(HistoricoLicitacion)
        .filter(HistoricoLicitacion.fecha_cierre >= limite_fecha)
        .order_by(HistoricoLicitacion.fecha_cierre.desc())
        .all()
    )
    historico_dict = [
        {
            "codigo_licitacion": h.codigo_licitacion or "",
            "cliente": h.cliente or "",
            "rubro": h.rubro or "",
            "monto_ofertado": h.monto_ofertado or 0.0,
            "margen_pct": h.margen_pct or 0.0,
            "fue_adjudicada": bool(h.fue_adjudicada),
            "fecha_cierre": h.fecha_cierre.isoformat() if h.fecha_cierre else "",
        }
        for h in historico
    ]

    atractividad = calcular_atractividad_licitacion(
        prediccion_actual=pred_actual,
        historico_3y=historico_dict,
        cliente_objetivo=proyecto.cliente or "",
    )

    return {
        "proyecto_id": proyecto.id,
        "codigo_licitacion": proyecto.codigo_licitacion or "",
        "cliente": proyecto.cliente or "",
        "atractividad": atractividad,
    }


@app.post("/api/prediccion/atractividad/ml/train")
async def entrenar_modelo_atractividad_ml(req: TrainMLAtractividadRequest, db: Session = Depends(get_db)):
    limite_fecha = datetime.utcnow() - timedelta(days=max(1, req.ultimos_anios) * 365)
    historico = (
        db.query(HistoricoLicitacion)
        .filter(HistoricoLicitacion.fecha_cierre >= limite_fecha)
        .order_by(HistoricoLicitacion.fecha_cierre.asc())
        .all()
    )
    historico_dict = [
        {
            "cliente": h.cliente or "",
            "rubro": h.rubro or "",
            "monto_ofertado": h.monto_ofertado or 0.0,
            "margen_pct": h.margen_pct or 0.0,
            "fue_adjudicada": bool(h.fue_adjudicada),
        }
        for h in historico
    ]
    trained = train_logistic_model(historico_dict, epochs=req.epochs, lr=req.lr)
    if not trained.get("ok"):
        raise HTTPException(status_code=400, detail=trained.get("error", "No se pudo entrenar modelo ML"))

    payload = {
        "version": "atractividad-ml-v2",
        "trained_at": datetime.utcnow().isoformat(),
        "config": {"ultimos_anios": req.ultimos_anios, "epochs": req.epochs, "lr": req.lr},
        "model": trained["model"],
    }
    ML_ATRACTIVIDAD_MODEL_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return {
        "message": "Modelo ML entrenado",
        "model_path": str(ML_ATRACTIVIDAD_MODEL_PATH.relative_to(BASE_DIR)),
        "metrics": {
            "train_size": trained["model"]["train_size"],
            "val_size": trained["model"]["val_size"],
            "accuracy_train": trained["model"]["accuracy_train"],
            "accuracy_val": trained["model"]["accuracy_val"],
        },
    }


@app.get("/api/prediccion/atractividad/ml/status")
async def estado_modelo_atractividad_ml():
    if not ML_ATRACTIVIDAD_MODEL_PATH.exists():
        return {"ready": False, "message": "Modelo ML no entrenado"}
    try:
        payload = json.loads(ML_ATRACTIVIDAD_MODEL_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {"ready": False, "message": "Modelo ML corrupto"}
    m = payload.get("model", {})
    return {
        "ready": True,
        "version": payload.get("version", ""),
        "trained_at": payload.get("trained_at", ""),
        "train_size": m.get("train_size", 0),
        "val_size": m.get("val_size", 0),
        "accuracy_train": m.get("accuracy_train", 0),
        "accuracy_val": m.get("accuracy_val", 0),
    }


@app.post("/api/prediccion/atractividad/ml/evaluar")
async def evaluar_atractividad_ml_v2(req: AtractividadMLV2Request, db: Session = Depends(get_db)):
    if not ML_ATRACTIVIDAD_MODEL_PATH.exists():
        raise HTTPException(status_code=400, detail="Modelo ML no entrenado. Ejecuta /api/prediccion/atractividad/ml/train")

    proyecto = db.query(Proyecto).filter(Proyecto.id == req.proyecto_id).first()
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")

    payload = json.loads(ML_ATRACTIVIDAD_MODEL_PATH.read_text(encoding="utf-8"))
    model = payload.get("model", {})
    cliente_eval = (req.cliente or "").strip() or (proyecto.cliente or "")
    rubro_eval = (req.rubro or "").strip()
    ml_result = predict_atractividad_ml(
        model=model,
        cliente=cliente_eval,
        rubro=rubro_eval,
        monto_oferta=req.monto_oferta or 0.0,
        margen_pct=req.margen_pct or 0.0,
    )
    return {
        "proyecto_id": proyecto.id,
        "codigo_licitacion": proyecto.codigo_licitacion or "",
        "cliente": cliente_eval,
        "rubro": rubro_eval,
        "resultado_ml": ml_result,
        "modelo": {
            "version": payload.get("version", ""),
            "trained_at": payload.get("trained_at", ""),
            "accuracy_val": model.get("accuracy_val", 0),
        },
    }


@app.post("/api/prediccion/bid-leveling")
async def bid_leveling(req: BidLevelingRequest, db: Session = Depends(get_db)):
    proyecto = db.query(Proyecto).filter(Proyecto.id == req.proyecto_id).first()
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    if len(req.oferta_ids or []) < 2:
        raise HTTPException(status_code=400, detail="Debes seleccionar al menos 2 ofertas")

    ofertas = (
        db.query(OfertaLicitacion)
        .filter(
            OfertaLicitacion.proyecto_id == req.proyecto_id,
            OfertaLicitacion.id.in_(req.oferta_ids),
        )
        .all()
    )
    if len(ofertas) < 2:
        raise HTTPException(status_code=400, detail="No se encontraron suficientes ofertas para comparar")

    min_monto = min((o.monto_oferta or 0.0) for o in ofertas) or 1.0
    plazos_validos = [o.plazo_dias for o in ofertas if (o.plazo_dias or 0) > 0]
    min_plazo = min(plazos_validos) if plazos_validos else 1

    comparativo = []
    for oferta in ofertas:
        pred = calcular_prediccion(proyecto.datos_extraidos or {}, oferta.factores or {})
        monto = oferta.monto_oferta or 0.0
        plazo = oferta.plazo_dias or 0
        precio_relativo = (monto / min_monto * 100.0) if min_monto > 0 else 100.0
        plazo_relativo = (plazo / min_plazo * 100.0) if plazo > 0 and min_plazo > 0 else 100.0

        comparativo.append({
            "oferta_id": oferta.id,
            "nombre": oferta.nombre,
            "monto_oferta": monto,
            "plazo_dias": plazo,
            "score": pred.get("score", 0.0),
            "score_base": pred.get("score_base", pred.get("score", 0.0)),
            "probabilidad": pred.get("probabilidad_adjudicacion", 0.0),
            "indice_riesgo": pred.get("indice_riesgo", 0.0),
            "precio_relativo_indice": round(precio_relativo, 2),
            "plazo_relativo_indice": round(plazo_relativo, 2),
            "top_riesgos": [r for r in (pred.get("matriz_riesgos", []) or []) if r.get("nivel") == "alto"],
            "prediccion": pred,
        })

    ranking = sorted(
        comparativo,
        key=lambda x: (
            x.get("score", 0.0),
            -(100.0 - min(200.0, x.get("precio_relativo_indice", 100.0))),
            -(100.0 - min(200.0, x.get("plazo_relativo_indice", 100.0))),
        ),
        reverse=True,
    )

    mejor = ranking[0]
    for item in ranking:
        item["brecha_score_vs_mejor"] = round(item["score"] - mejor["score"], 2)
        item["brecha_precio_vs_mejor_pct"] = round(item["precio_relativo_indice"] - mejor["precio_relativo_indice"], 2)
        item["brecha_plazo_vs_mejor_pct"] = round(item["plazo_relativo_indice"] - mejor["plazo_relativo_indice"], 2)
        item["brecha_riesgo_vs_mejor"] = round(item["indice_riesgo"] - mejor["indice_riesgo"], 2)

    resumen = {
        "oferta_recomendada_id": mejor["oferta_id"],
        "oferta_recomendada_nombre": mejor["nombre"],
        "justificacion": (
            f"Mayor score ajustado por riesgo ({mejor['score']:.2f}) "
            f"y probabilidad estimada ({mejor['probabilidad']*100:.2f}%)."
        ),
    }

    return {
        "proyecto_id": req.proyecto_id,
        "cantidad_ofertas": len(ranking),
        "ranking": ranking,
        "resumen": resumen,
    }

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("backend.main:app", host="0.0.0.0", port=8000, reload=True)
